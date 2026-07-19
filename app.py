import streamlit as st
import pandas as pd
import numpy as np
from io import StringIO, BytesIO
import json
import os
import base64
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="Reconciliation Dashboard Partner",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# LOGOS (embedded base64)
# ============================================================
def load_logo(path, mime):
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), path),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), os.path.basename(path)),
        path,
    ]
    for p in candidates:
        try:
            with open(p, 'rb') as f:
                b64 = base64.b64encode(f.read()).decode()
            return f"data:image/{mime};base64,{b64}"
        except:
            continue
    return ""

LOGO_PAYX    = load_logo("logos/payx_logo.svg", "svg+xml")
LOGO_PARTNER = load_logo("logos/logo_partner_internet.png", "png")
LOGO_012     = load_logo("logos/talk012_logo.png", "png")
LOGO_PELE    = load_logo("logos/pelephoen.png", "png")
LOGO_CELL    = load_logo("logos/cellcom.png", "png")

# ============================================================
# STYLES
# ============================================================
st.markdown("""
<style>
    #MainMenu {visibility: hidden;}
    .main-header {
        background: linear-gradient(135deg, #1F3864, #2E75B6);
        color: white;
        padding: 16px 24px;
        border-radius: 10px;
        margin-bottom: 20px;
        display: flex;
        align-items: center;
        gap: 20px;
    }
    .main-header h1 { margin: 0; font-size: 24px; }
    .main-header p  { margin: 4px 0 0 0; opacity: 0.85; font-size: 13px; }
    .logo-row { display: flex; align-items: center; gap: 16px; margin-bottom: 16px; }
    .logo-row img { height: 36px; object-fit: contain; }
    .action-box {
        background: #1a1f2e;
        border: 1px solid #e74c3c;
        border-radius: 8px;
        padding: 14px 18px;
        margin-bottom: 16px;
    }
    .action-box h4 { color: #e74c3c; margin: 0 0 10px 0; font-size: 14px; }
    .stDataFrame { font-size: 12px; }
    div[data-testid="stMetricValue"] { font-size: 26px; font-weight: bold; }
    .sheets-error {
        background: #fff3cd;
        border: 2px solid #e74c3c;
        border-radius: 8px;
        padding: 12px 16px;
        margin-bottom: 16px;
        color: #721c24;
        font-weight: 500;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================
# CELLCOM PRICE MAP
# ============================================================
APP_VERSION = "v2026-07-19-ravkav-fix"

CELLCOM_FIXED = {15.0, 19.0, 25.0, 29.0, 39.9, 49.0}
CELLCOM_DISCOUNT = 5.0

def cellcom_expected_supplier_price(our_eup):
    if our_eup in CELLCOM_FIXED:
        return our_eup
    return round(our_eup - CELLCOM_DISCOUNT, 2)

# ============================================================
# GOOGLE SHEETS
# ============================================================
SCOPES = ['https://www.googleapis.com/auth/spreadsheets',
          'https://www.googleapis.com/auth/drive']

HISTORY_COLS = ['date','operator_tab','sup_cbd','our_eup','diff',
                'matched_count','sup_only_count','sup_only_cbd',
                'our_only_count','our_only_eup','real_gap',
                'pending_count','refunds_eup','net_billed',
                'esim_count','esim_our_total','esim_sup_total']

DETAIL_COLS = ['date','operator_tab','category','phone','operator',
               'product','amount','sup_date','our_date','reason','check_instruction','verified',
               'verified_by','verified_at']

# ---- FIX 1: Sheets connectivity check with visible banner ----
@st.cache_resource
def get_gspread_client():
    """Returns gspread client — cached for entire app lifetime."""
    if not GSPREAD_AVAILABLE:
        return None
    import time
    for attempt in range(3):
        try:
            info = dict(st.secrets["gcp_service_account"])
            if 'private_key' in info:
                info['private_key'] = info['private_key'].replace('\\n', '\n')
            creds = Credentials.from_service_account_info(info, scopes=SCOPES)
            gc = gspread.authorize(creds)
            return gc
        except Exception:
            if attempt < 2:
                time.sleep(1)
    return None

def check_sheets_banner():
    """Call at the top of any save operation — shows prominent warning if Sheets is down."""
    if get_spreadsheet("partner") is None:
        st.markdown("""
        <div class="sheets-error">
        ⛔ <strong>Google Sheets unavailable.</strong> Data will be saved LOCALLY
        and may be lost on server restart. Contact your administrator
        to restore the connection before saving.
        </div>
        """, unsafe_allow_html=True)
        return False
    return True

def check_existing_record(report_date, operator_tab):
    """
    Returns True if a history record for this date+operator already exists in Sheets.
    Used to warn the user before saving a second time for the same day.
    """
    sh = get_spreadsheet(operator_tab)
    if sh is None:
        return False
    try:
        dt = datetime.strptime(report_date, '%Y-%m-%d')
        ws_title = dt.strftime('%B %Y')
        try:
            ws = sh.worksheet(ws_title)
        except gspread.WorksheetNotFound:
            return False
        records = ws.get_all_records()
        for r in records:
            if (str(r.get('date', '')) == str(report_date) and
                    str(r.get('operator_tab', '')) == operator_tab):
                return True
        return False
    except Exception:
        return False

def save_confirmation_ui(report_date, operator_tab, session_key):
    """
    Shows a confirmation dialog if a record for this date already exists.
    Returns True if user confirmed (or no existing record).
    Returns False if user cancelled or hasn't confirmed yet.

    Uses session_state to track the confirmation step:
      session_key + '_confirm_pending' = True  → waiting for user choice
      session_key + '_confirm_ok'      = True  → user said yes, proceed
    """
    confirm_pending_key = session_key + '_confirm_pending'
    confirm_ok_key      = session_key + '_confirm_ok'

    # Already confirmed in this click cycle → proceed
    if st.session_state.get(confirm_ok_key):
        st.session_state.pop(confirm_ok_key, None)
        st.session_state.pop(confirm_pending_key, None)
        return True

    # Check if a record already exists
    if not st.session_state.get(confirm_pending_key):
        if check_existing_record(report_date, operator_tab):
            st.session_state[confirm_pending_key] = True
        else:
            return True  # No existing record — safe to save, no confirmation needed

    # Show the confirmation dialog
    if st.session_state.get(confirm_pending_key):
        st.warning(
            f"⚠️ Record for **{report_date}** ({operator_tab}) already exists in Google Sheets.\n\n"
            f"What would you like to do?"
        )
        col_yes, col_no = st.columns(2)
        with col_yes:
            if st.button("✅ Save as new version", key=session_key + '_yes', use_container_width=True):
                st.session_state[confirm_ok_key] = True
                st.session_state.pop(confirm_pending_key, None)
                st.rerun()
        with col_no:
            if st.button("❌ Cancel", key=session_key + '_no', use_container_width=True):
                st.session_state.pop(confirm_pending_key, None)
                st.session_state['show_cancel_msg'] = True
        if st.session_state.get('show_cancel_msg'):
            st.session_state.pop('show_cancel_msg', None)
            st.info("Save cancelled. Existing data unchanged.")
            return False
        return False  # Wait for user choice

    return True

@st.cache_resource(ttl=1800, show_spinner=False)
def _open_spreadsheet(operator):
    """Open spreadsheet once per operator, cached 30 min.
    Raises on failure so that failures are NOT cached."""
    gc = get_gspread_client()
    if gc is None:
        raise RuntimeError("gspread client unavailable")
    if operator == 'pelephone':
        sid = st.secrets["google_sheets_pelephone"]["spreadsheet_id"]
    elif operator == 'ravkav':
        sid = st.secrets["google_sheets_ravkav"]["spreadsheet_id"]
    elif operator == 'cellcom':
        sid = st.secrets["google_sheets_cellcom"]["spreadsheet_id"]
    else:
        sid = st.secrets["google_sheets"]["spreadsheet_id"]
    return gc.open_by_key(sid)

def get_spreadsheet(operator='partner'):
    try:
        return _open_spreadsheet(operator)
    except Exception:
        return None

def get_or_create_sheet(sh, title, headers):
    try:
        ws = sh.worksheet(title)
        existing = [c for c in ws.row_values(1) if c]
        if not existing:
            ws.append_row(headers)
        return ws
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=2000, cols=len(headers))
        ws.append_row(headers)
        return ws

def _is_month_sheet(title):
    months = ['January','February','March','April','May','June',
              'July','August','September','October','November','December']
    return any(m in title for m in months)

@st.cache_data(ttl=30, show_spinner=False)
def load_history(month=None, operator_tab=None):
    sh = get_spreadsheet(operator_tab or 'partner')
    if sh is None: return _load_local_history()
    try:
        if month:
            dt = datetime.strptime(month, '%Y-%m')
            ws = get_or_create_sheet(sh, dt.strftime('%B %Y'), HISTORY_COLS)
            _ensure_history_headers(ws)
            records = ws.get_all_records()
        else:
            records = []
            for ws in sh.worksheets():
                if _is_month_sheet(ws.title):
                    try: records.extend(ws.get_all_records())
                    except: pass
            records = sorted(records, key=lambda x: x.get('date',''))
        for r in records:
            if not r.get('operator_tab'):
                r['operator_tab'] = 'partner'
        if operator_tab:
            records = [r for r in records if r.get('operator_tab','') == operator_tab]
        return records
    except Exception:
        return _load_local_history()

# ---- FIX 2: safe save_to_sheets — NEVER overwrites, always appends ----
def save_to_sheets(record):
    """
    SAFE version: always appends a new row.
    If a row for the same date+operator already exists, it is LEFT INTACT.
    The new row is added below it. Both rows remain in the sheet permanently.
    To see the latest result, look at the last row for that date.
    """
    sh = get_spreadsheet(record.get('operator_tab', 'partner'))
    if sh is None:
        _save_local_history(record)
        return False, "⚠️ Sheets unavailable — saved locally (risk of loss on restart!)"
    try:
        dt = datetime.strptime(record['date'], '%Y-%m-%d')
        ws = get_or_create_sheet(sh, dt.strftime('%B %Y'), HISTORY_COLS)
        _ensure_history_headers(ws)

        # SAFE: always append, never update existing rows
        ws.append_row([record.get(c, '') for c in HISTORY_COLS])
        return True, f"New record added for {record['date']} in '{dt.strftime('%B %Y')}'"
    except Exception as e:
        _save_local_history(record)
        return False, f"⚠️ Sheets error: {e} — saved locally"

# ---- FIX 3: safe save_details_to_sheets — NO ws.clear(), uses append only ----
def save_details_to_sheets(report_date, operator_tab, rows):
    """
    SAFE version: does NOT clear the sheet.
    Checks for existing rows for this date+operator:
    - If verified rows exist → they are preserved, only unverified rows are replaced
    - New rows are appended for this date
    This means: verified statuses are NEVER lost on re-run.
    """
    sh = get_spreadsheet(operator_tab)
    if sh is None: return False, "Not connected"
    try:
        ws = get_or_create_sheet(sh, 'Transaction Details', DETAIL_COLS)
        _ensure_detail_headers(ws)
        all_data = ws.get_all_records()

        # Split: keep verified rows for this date (protect them), drop unverified for this date
        verified_this_date = [
            r for r in all_data
            if str(r.get('date', '')) == str(report_date)
            and str(r.get('operator_tab', '')) == operator_tab
            and not str(r.get('verified', '')).startswith('⬜')  # Keep: already verified
        ]
        other_dates = [
            r for r in all_data
            if not (str(r.get('date', '')) == str(report_date)
                    and str(r.get('operator_tab', '')) == operator_tab)
        ]

        # Rebuild: other dates + verified from this date + new unverified rows
        rebuilt = other_dates + verified_this_date

        # SAFE: prepare ALL rows before touching the sheet.
        # If row-building fails here, ws.clear() is never called → data stays intact.
        rebuilt_rows = [[r.get(c, '') for c in DETAIL_COLS] for r in rebuilt]
        new_rows = rows if rows else []
        all_rows_to_write = rebuilt_rows + new_rows

        # Now clear and rewrite in one batch — minimises the window where sheet is empty
        ws.clear()
        ws.append_row(DETAIL_COLS)
        if all_rows_to_write:
            ws.append_rows(all_rows_to_write)

        preserved = len(verified_this_date)
        added = len(rows)
        msg = f"Added {added} rows"
        if preserved > 0:
            msg += f" | Preserved {preserved} verified records"
        return True, msg
    except Exception as e:
        return False, f"Details error: {e}"

def load_pending_verifications():
    """One read per operator; every pending record carries its real sheet
    row number (_row) and source spreadsheet (_op) for direct writes."""
    all_records = []
    errors = []
    for op in ['partner', 'pelephone', 'cellcom', 'ravkav']:
        sh = get_spreadsheet(op)
        if sh is None:
            errors.append(f"{op}: not connected")
            continue
        try:
            ws = get_or_create_sheet(sh, 'Transaction Details', DETAIL_COLS)
            _ensure_detail_headers(ws)
            values = _api_retry(lambda: ws.get_all_values())
            if not values:
                continue
            headers = values[0]
            for idx, rowvals in enumerate(values[1:], start=2):
                r = dict(zip(headers, rowvals))
                if str(r.get('verified', '')).startswith('⬜'):
                    r['_row'] = idx   # real row number in the sheet
                    r['_op'] = op     # which spreadsheet it lives in
                    all_records.append(r)
        except Exception as e:
            errors.append(f"{op}: {str(e)[:60]}")
    if errors:
        st.session_state['pending_load_errors'] = errors
    else:
        st.session_state.pop('pending_load_errors', None)
    return all_records

@st.cache_data(ttl=120, show_spinner=False)
def load_month_details(operator_tab, month):
    """All Transaction Details rows of one operator for YYYY-MM (cached 2 min)."""
    sh = get_spreadsheet(operator_tab)
    if sh is None:
        return []
    try:
        ws = sh.worksheet('Transaction Details')
        _ensure_detail_headers(ws)
        recs = _api_retry(lambda: ws.get_all_records())
        return [r for r in recs
                if str(r.get('operator_tab', '')) == operator_tab
                and str(r.get('date', '')).startswith(month)]
    except Exception:
        return []

def load_verified():
    all_records = []
    for op in ['partner', 'pelephone', 'cellcom', 'ravkav']:
        sh = get_spreadsheet(op)
        if sh is None: continue
        try:
            ws = get_or_create_sheet(sh, 'Transaction Details', DETAIL_COLS)
            _ensure_detail_headers(ws)
            records = ws.get_all_records()
            all_records.extend([r for r in records
                                if str(r.get('verified','')).startswith('✅')
                                or str(r.get('verified','')).startswith('❌')
                                or str(r.get('verified','')).startswith('🔵')])
        except: pass
    return all_records

def _api_retry(fn, attempts=3):
    """Run a Sheets API call with backoff retry on quota/server errors (429/5xx)."""
    import time
    last_err = None
    for a in range(attempts):
        try:
            return fn()
        except gspread.exceptions.APIError as e:
            last_err = e
            code = getattr(getattr(e, 'response', None), 'status_code', None)
            if code in (429, 500, 502, 503) and a < attempts - 1:
                time.sleep(2 ** (a + 1))  # 2s, 4s
                continue
            raise
    raise last_err

VERIFIED_COL = DETAIL_COLS.index('verified') + 1
PHONE_COL = DETAIL_COLS.index('phone') + 1
VERIFIED_BY_COL = DETAIL_COLS.index('verified_by') + 1
VERIFIED_AT_COL = DETAIL_COLS.index('verified_at') + 1

def _current_user():
    return st.session_state.get('auth_user', '') or 'unknown'

def _audit_cells(row_num, status):
    """Status + who + when for one sheet row (single batch write)."""
    return [gspread.Cell(row=row_num, col=VERIFIED_COL, value=status),
            gspread.Cell(row=row_num, col=VERIFIED_BY_COL, value=_current_user()),
            gspread.Cell(row=row_num, col=VERIFIED_AT_COL, value=_now_il())]

def _ensure_history_headers(ws):
    """Lazy migration of month-sheet headers (adds eSIM columns once)."""
    try:
        hdr = ws.row_values(1)
        if 'esim_count' not in hdr:
            rng = f"A1:{get_column_letter(len(HISTORY_COLS))}1"
            _api_retry(lambda: ws.batch_update(
                [{'range': rng, 'values': [HISTORY_COLS]}]))
    except Exception:
        pass

def _ensure_detail_headers(ws):
    """One-time lazy migration: extend the header row of old 12-column
    'Transaction Details' sheets so audit values survive daily rebuilds."""
    try:
        hdr = ws.row_values(1)
        if 'verified_by' not in hdr:
            rng = f"A1:{get_column_letter(len(DETAIL_COLS))}1"
            _api_retry(lambda: ws.batch_update(
                [{'range': rng, 'values': [DETAIL_COLS]}]))
    except Exception:
        pass

def _norm_cmp(p):
    return str(p or '').strip().replace('.0', '').lstrip('0')

def update_verification_by_row(sh, sheet_row, phone, new_status):
    """Direct row write: 1 cheap read (integrity check of one cell) + 1 write.
    No full-sheet read -> no 429 cascade. If the row shifted (daily save
    rebuilt the sheet), refuse to write and ask for a Refresh."""
    try:
        if not sheet_row:
            return False, "No row reference — click 🔄 Refresh to reload."
        ws = sh.worksheet('Transaction Details')
        current_phone = _api_retry(lambda: ws.cell(sheet_row, PHONE_COL).value)
        if _norm_cmp(current_phone) != _norm_cmp(phone):
            return False, (f"Row {sheet_row} changed (holds '{current_phone}', "
                           f"expected '{phone}'). Click 🔄 Refresh and retry.")
        _api_retry(lambda: ws.update_cells(
            _audit_cells(sheet_row, new_status), value_input_option='RAW'))
        return True, f"Updated {phone} (row {sheet_row})"
    except Exception as e:
        return False, f"Sheets error: {e}"

def find_date_shift_pairs(records, max_shift_days=4):
    """Partner only: pair 'Our Only' (date D) with 'Supplier Only' (D+1..D+max_shift_days)
    for the same normalized phone and identical amount. One-to-one matching,
    earliest supplier date wins. Returns list of (our_rec, sup_rec, shift_days)."""
    from datetime import datetime as _dt

    def _pdate(s):
        try:
            return _dt.strptime(str(s).strip()[:10], '%Y-%m-%d')
        except Exception:
            return None

    def _amt(v):
        try:
            return round(float(str(v).replace(',', '')), 2)
        except Exception:
            return None

    ours, sups = [], []
    for r in records:
        if r.get('_op', r.get('operator_tab')) != 'partner':
            continue
        cat = str(r.get('category', '')).strip()
        d = _pdate(r.get('date'))
        a = _amt(r.get('amount'))
        if d is None or a is None:
            continue
        key = (_norm_cmp(r.get('phone')), a)
        if cat == 'Our Only':
            ours.append((key, d, r))
        elif cat == 'Supplier Only':
            sups.append((key, d, r))

    sups.sort(key=lambda x: x[1])
    used = set()
    pairs = []
    for key_o, d_o, r_o in sorted(ours, key=lambda x: x[1]):
        for j, (key_s, d_s, r_s) in enumerate(sups):
            if j in used or key_s != key_o:
                continue
            delta = (d_s - d_o).days
            if 1 <= delta <= max_shift_days:
                used.add(j)
                pairs.append((r_o, r_s, delta))
                break
    return pairs

def cross_day_match(result, report_date, operator_tab):
    sh = get_spreadsheet(operator_tab)
    if sh is None: return [], []
    try:
        ws = sh.worksheet('Transaction Details')
        all_details = ws.get_all_records()
        df = pd.DataFrame(all_details)
        if df.empty: return [], []
        df = df[df['operator_tab'] == operator_tab]
        dt = datetime.strptime(report_date, '%Y-%m-%d')
        yesterday = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
        yest_sup = set(df[(df['date']==yesterday) & (df['category']=='Supplier Only')]['phone'])
        yest_our = set(df[(df['date']==yesterday) & (df['category']=='Our Only')]['phone'])
        our_only  = set(result['our_only']['Phone_Display']) if len(result['our_only']) > 0 else set()
        sup_only  = set(result['sup_only']['Phone_Display']) if len(result['sup_only']) > 0 else set()
        return list(our_only & yest_sup), list(sup_only & yest_our)
    except: return [], []

# ---- FIX 4: local fallback — also safe (append semantics) ----
HISTORY_FILE = os.path.join(os.path.dirname(__file__), "history.json")

def _load_local_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE,'r',encoding='utf-8') as f: return json.load(f)
    return []

def _save_local_history(record):
    """
    SAFE local fallback: appends the new record without deleting any existing records.
    Multiple records for the same date are allowed — last one is the most recent.
    """
    h = _load_local_history()
    # Append only — do NOT filter/delete existing records for the same date
    h.append(record)
    h.sort(key=lambda x: (x.get('date',''), x.get('operator_tab','')))
    with open(HISTORY_FILE,'w',encoding='utf-8') as f:
        json.dump(h, f, ensure_ascii=False, indent=2)

# ============================================================
# PHONE NORMALIZATION
# ============================================================
def norm_phone(phone):
    if phone is None or (isinstance(phone, float) and np.isnan(phone)): return ''
    s = str(phone).strip().replace('.0','').replace(' ','').replace('+','')
    if 'E' in s.upper():
        try: s = str(int(float(s)))
        except: pass
    s = s.replace('.0','')
    if s.startswith('00972'): s = s[5:]
    elif s.startswith('972'): s = s[3:]
    if s.startswith('0'): s = s[1:]
    return s.strip()

def display_phone(norm):
    if norm and len(norm) >= 8: return '0' + norm
    return norm

# ============================================================
# LOAD OUR FILES
# ============================================================
def load_our(file_bytes, operator_name):
    try:
        for enc in ['utf-8-sig','utf-8','windows-1255','cp1255','latin1']:
            try: text = file_bytes.decode(enc); break
            except: continue
        df = pd.read_csv(StringIO(text), dtype={'Phone Number': str}, on_bad_lines='skip')
        df['Operator'] = operator_name
        df['End User Price'] = pd.to_numeric(df.get('End User Price',0), errors='coerce').fillna(0)
        df['Customer price'] = pd.to_numeric(df.get('Customer price',0), errors='coerce').fillna(0)
        df = df[~df['Action'].isin(['REWARD','REFUND_REWARD'])]
        df['Is_Refund'] = (df['Action'] == 'REFUND') & (df['Status'] == 'DONE')
        df['Eff_Status'] = df.apply(
            lambda r: 'CANCELLED' if (r['Action']=='REFUND' and r['Status']=='DONE') else r['Status'], axis=1)
        df['phone_norm'] = df['Phone Number'].apply(norm_phone)
        df['Phone_Display'] = df['phone_norm'].apply(display_phone)
        try:
            df['_dt'] = pd.to_datetime(df['Date & Time'], dayfirst=True, errors='coerce')
            df['Is_Late'] = df['_dt'].dt.hour >= 22
        except: df['Is_Late'] = False
        df = df[df['phone_norm'].str.len() >= 7]
        return df, None
    except Exception as e: return None, str(e)

# ============================================================
# LOAD SUPPLIER FILES
# ============================================================
def load_supplier_partner(file_bytes, filename):
    try:
        if filename.lower().endswith('.xls'):
            text = file_bytes.decode('cp1255', errors='replace')
        else:
            for enc in ['utf-8-sig','windows-1255','cp1255','latin1']:
                try: text = file_bytes.decode(enc); break
                except: continue
        lines = text.replace('\r\n','\n').replace('\r','\n').strip().split('\n')
        start = 1 if len([c for c in lines[0].split('\t') if c.strip()]) == 1 else 0
        sep = '\t' if lines[start].count('\t') > lines[start].count(',') else ','
        df = pd.read_csv(StringIO('\n'.join(lines[start:])), sep=sep,
                         dtype={'MSISDN': str}, on_bad_lines='skip')
        col_map = {
            'שם לקוח':'Cust_Name','מספר לקוח':'Cust_Num',
            'מס טרנזקציה':'Sup_TxID','תאריך ושעת טעינה':'Sup_Date',
            'MSISDN':'MSISDN','חיוב לפני הנחה כולל מעמ':'CBD',
            'נטו כולל מעמ':'Net_Total','שם כרטיס':'Package',
            'סוג כרטיס':'Card_Type','אחוז הנחה':'Discount_Pct',
            'סכום הנחה כולל מעמ':'Discount_Amt'
        }
        df.rename(columns=col_map, inplace=True)
        df['CBD'] = pd.to_numeric(df.get('CBD',0), errors='coerce').fillna(0)
        df['Net_Total'] = pd.to_numeric(df.get('Net_Total',0), errors='coerce').fillna(0)
        df['phone_norm'] = df['MSISDN'].apply(norm_phone)
        df['Phone_Display'] = df['phone_norm'].apply(display_phone)
        df = df[df['phone_norm'].str.len() >= 8]
        return df, None
    except Exception as e: return None, str(e)

def load_supplier_cellcom(file_bytes):
    try:
        df = pd.read_excel(BytesIO(file_bytes))
        col_map = {
            'שם משווק':'Dealer_Name',
            'תאריך טעינה':'Sup_Date',
            'סכום הטענה':'CBD',
            'סכום ביטול':'Cancel_Amt',
            'קוד כרטיס':'Card_Code',
            'טרמינל':'Terminal',
            'מספר מנוי':'MSISDN',
            'BAN':'BAN'
        }
        df.rename(columns=col_map, inplace=True)
        df['CBD'] = pd.to_numeric(df.get('CBD',0), errors='coerce').fillna(0)
        df['Cancel_Amt'] = pd.to_numeric(df.get('Cancel_Amt',0), errors='coerce').fillna(0)
        df['Is_Cancel'] = df['Cancel_Amt'].notna() & (df['Cancel_Amt'] > 0)
        df['phone_norm'] = df['MSISDN'].apply(norm_phone)
        df['Phone_Display'] = df['phone_norm'].apply(display_phone)
        df['Sup_Date'] = pd.to_datetime(df['Sup_Date'], errors='coerce')
        df = df[df['phone_norm'].str.len() >= 8]
        df = df[df['CBD'] > 0]
        return df, None
    except Exception as e: return None, str(e)

def load_supplier_pelephone(file_bytes):
    try:
        df = pd.read_excel(BytesIO(file_bytes))
        rename_map = {
            'Serial_Number':'Serial',
            "#DOC_NUMBER'":'Doc_Number',
            'Order_Number':'Order_Number',
            'TOPUP_TIME':'Sup_Time',
            'TOPUP_DATE':'Sup_Date',
            'dealer_price':'Dealer_Price',
            'TOPUP_ITEM':'TOPUP_ITEM',
            'Dealer':'Dealer',
            'SUBSCRIBER':'MSISDN',
        }
        if 'Unnamed: 5' in df.columns:
            rename_map['Unnamed: 5'] = 'TOPUP_PRICE'
        unnamed_cols = [c for c in df.columns if str(c).startswith('Unnamed')]
        if 'MSISDN' not in df.columns and 'SUBSCRIBER' not in df.columns and unnamed_cols:
            rename_map[unnamed_cols[-1]] = 'MSISDN'
        df.rename(columns=rename_map, inplace=True)
        df['TOPUP_PRICE'] = pd.to_numeric(df.get('TOPUP_PRICE',0), errors='coerce').fillna(0)
        df['phone_norm'] = df['MSISDN'].apply(norm_phone)
        df['Phone_Display'] = df['phone_norm'].apply(display_phone)
        df['Order_Number'] = df['Order_Number'].astype(str).str.strip()
        try:
            df['Sup_DateTime'] = pd.to_datetime(
                df['Sup_Date'].astype(str) + ' ' + df['Sup_Time'].astype(str),
                dayfirst=True, errors='coerce')
        except: df['Sup_DateTime'] = pd.NaT
        df = df[df['phone_norm'].str.len() >= 8]
        return df, None
    except Exception as e: return None, str(e)

# ============================================================
# CHECK INSTRUCTIONS
# ============================================================
def make_check_instruction(category, sup_date, our_date, report_date, is_late=False):
    rd = str(report_date)
    if category == 'Supplier Only':
        try:
            if hasattr(sup_date, 'strftime'):
                sd = sup_date
            else:
                s = str(sup_date).strip()
                try: sd = pd.to_datetime(s, dayfirst=False)
                except: sd = pd.to_datetime(s, dayfirst=True)
            sd_str = sd.strftime('%d-%b-%Y')
            if sd.strftime('%Y-%m-%d') != rd:
                return f"Check OUR reports for {sd_str} — verify this phone appears there. If yes → date shift, not a real gap."
            else:
                return "Transaction date matches report date. Check if it was processed in our system."
        except:
            return "Check our system for this transaction date."
    elif category == 'Our Only':
        try:
            if hasattr(our_date, 'strftime'):
                dt = our_date
            else:
                dt = pd.to_datetime(str(our_date), dayfirst=True)
            next_day = (dt + timedelta(days=1)).strftime('%d-%b-%Y')
            if is_late:
                return f"Late transaction ({dt.strftime('%H:%M')}). Check SUPPLIER report for {next_day} — likely appears there."
            else:
                return f"Check SUPPLIER report for {next_day} — may have been processed next day."
        except:
            return "Check supplier report for next day."
    return ""


def run_recon_partner(sup_df, partner_df, talk_df, report_date):
    our_all = pd.concat([partner_df, talk_df], ignore_index=True)
    our_dc      = our_all[(our_all['Eff_Status'].isin(['DONE','CANCELLED'])) & (~our_all['Is_Refund'])].copy()
    our_pending = our_all[our_all['Eff_Status'] == 'PENDING_CANCELLATION'].copy()
    our_refunds = our_all[our_all['Is_Refund']].copy()
    our_failed  = our_all[our_all['Eff_Status'] == 'FAILED'].copy()

    sup_phones    = set(sup_df['phone_norm'])
    our_dc_phones = set(our_dc['phone_norm'])
    our_pnd_phones= set(our_pending['phone_norm'])

    matched_phones  = sup_phones & our_dc_phones
    sup_only_phones = sup_phones - our_dc_phones - our_pnd_phones
    sup_pnd_phones  = sup_phones & our_pnd_phones
    our_only_phones = our_dc_phones - sup_phones

    matched_rows = []
    used_our = set()
    for _, sr in sup_df[sup_df['phone_norm'].isin(matched_phones)].iterrows():
        om = our_dc[(our_dc['phone_norm']==sr['phone_norm']) & (~our_dc['Transaction ID'].isin(used_our))]
        if len(om) > 0:
            or_ = om.iloc[0]
            used_our.add(or_['Transaction ID'])
            matched_rows.append({
                'Phone': or_['Phone_Display'],
                'Supplier Date': sr.get('Sup_Date',''),
                'Supplier Package': sr.get('Package',''),
                'Supplier Tx ID': sr.get('Sup_TxID',''),
                'Supplier CBD (NIS)': sr['CBD'],
                'Our Tx ID': or_['Transaction ID'],
                'Our Date': or_['Date & Time'],
                'Our Operator': or_['Operator'],
                'Our Status': or_['Eff_Status'],
                'Our Product': or_['Product Name'],
                'Our EUP (NIS)': or_['End User Price'],
                'Difference (NIS)': sr['CBD'] - or_['End User Price'],
            })

    matched_df = pd.DataFrame(matched_rows)
    sup_pure = sup_df.copy()
    sup_only_df = sup_pure[sup_pure['phone_norm'].isin(sup_only_phones)].copy()
    sup_only_df['Reason'] = 'Not found in our system'
    sup_only_df['Check_Instruction'] = sup_only_df.apply(
        lambda r: make_check_instruction('Supplier Only', r.get('Sup_Date',''), '', report_date), axis=1)

    sup_pnd_df = sup_df[sup_df['phone_norm'].isin(sup_pnd_phones)].copy()
    our_only_df = our_dc[our_dc['phone_norm'].isin(our_only_phones)].copy()
    our_only_df['Reason'] = our_only_df.apply(
        lambda r: 'Late transaction (after 22:00)' if r.get('Is_Late', False) else 'Not found at supplier', axis=1)
    our_only_df['Check_Instruction'] = our_only_df.apply(
        lambda r: make_check_instruction('Our Only', '', r.get('Date & Time',''), report_date, r.get('Is_Late',False)), axis=1)

    t = {
        'sup_cbd': matched_df['Supplier CBD (NIS)'].sum() if len(matched_df) else 0,
        'our_eup': matched_df['Our EUP (NIS)'].sum() if len(matched_df) else 0,
        'diff': matched_df['Difference (NIS)'].sum() if len(matched_df) else 0,
        'sup_only_cbd': sup_only_df['CBD'].sum() if len(sup_only_df) else 0,
        'sup_pnd_cbd': sup_pnd_df['CBD'].sum() if len(sup_pnd_df) else 0,
        'our_only_eup': our_only_df['End User Price'].sum() if len(our_only_df) else 0,
        'pending_eup': our_pending['End User Price'].sum() if len(our_pending) else 0,
        'refunds_eup': our_refunds['End User Price'].sum() if len(our_refunds) else 0,
        'partner_eup': partner_df[partner_df['Eff_Status'].isin(['DONE','CANCELLED']) & ~partner_df['Is_Refund']]['End User Price'].sum(),
        'talk012_eup': talk_df[talk_df['Eff_Status'].isin(['DONE','CANCELLED']) & ~talk_df['Is_Refund']]['End User Price'].sum(),
        'partner_ref': partner_df[partner_df['Is_Refund']]['End User Price'].sum(),
        'talk012_ref': talk_df[talk_df['Is_Refund']]['End User Price'].sum(),
        'matched_count': len(matched_phones),
        'sup_only_count': len(sup_only_phones),
        'sup_pnd_count': len(sup_pnd_phones),
        'our_only_count': len(our_only_phones),
        'pending_count': len(our_pending),
        'refunds_count': len(our_refunds),
        'failed_count': len(our_failed),
    }
    t['real_gap'] = round(t['sup_only_cbd'] - t['our_only_eup'], 2)

    return {'matched': matched_df, 'sup_only': sup_only_df, 'sup_pending': sup_pnd_df,
            'our_only': our_only_df, 'pending': our_pending, 'refunds': our_refunds,
            'failed': our_failed, 'totals': t}

# ============================================================
# RECONCILIATION — CELLCOM
# ============================================================
def run_recon_cellcom(sup_df, our_df, report_date):
    our_dc      = our_df[(our_df['Eff_Status'].isin(['DONE','CANCELLED'])) & (~our_df['Is_Refund'])].copy()
    our_refunds = our_df[our_df['Is_Refund']].copy()
    our_failed  = our_df[our_df['Eff_Status'] == 'FAILED'].copy()
    our_pending = our_df[our_df['Eff_Status'] == 'PENDING'].copy()
    sup_pure       = sup_df[(sup_df['CBD'] > 0) & (sup_df['Cancel_Amt'] == 0)].copy()
    sup_refunds_df = sup_df[(sup_df['CBD'] > 0) & (sup_df['Cancel_Amt'] > 0)].copy()

    sup_phones    = set(sup_pure['phone_norm'])
    our_dc_phones = set(our_dc['phone_norm'])
    matched_phones  = sup_phones & our_dc_phones
    sup_only_phones = sup_phones - our_dc_phones
    our_only_phones = our_dc_phones - sup_phones

    used_our_ref = set()
    used_sup_ref = set()
    unmatched_our_ref = []
    for si, sr in sup_refunds_df.iterrows():
        if si in used_sup_ref: continue
        om = our_refunds[(our_refunds['phone_norm'] == sr['phone_norm']) & (~our_refunds['Transaction ID'].isin(used_our_ref))]
        if len(om) > 0:
            used_our_ref.add(om.iloc[0]['Transaction ID'])
            used_sup_ref.add(si)
    for _, or_ in our_refunds.iterrows():
        if or_['Transaction ID'] not in used_our_ref:
            unmatched_our_ref.append(or_)
    unmatched_our_ref_df = pd.DataFrame(unmatched_our_ref) if unmatched_our_ref else pd.DataFrame()

    matched_rows = []
    used_our = set()
    used_sup = set()
    price_diffs = []

    sup_matched = sup_pure[sup_pure['phone_norm'].isin(matched_phones)].copy()

    for si, sr in sup_matched.iterrows():
        if si in used_sup: continue
        phone = sr['phone_norm']
        actual_sup = sr['CBD']
        om = our_dc[
            (our_dc['phone_norm'] == phone) &
            (~our_dc['Transaction ID'].isin(used_our)) &
            (our_dc['End User Price'].apply(cellcom_expected_supplier_price) == actual_sup)
        ]
        if len(om) > 0:
            or_ = om.iloc[0]
            used_our.add(or_['Transaction ID'])
            used_sup.add(si)
            our_eup = or_['End User Price']
            expected_sup = cellcom_expected_supplier_price(our_eup)
            price_diff = round(actual_sup - expected_sup, 2)
            price_diffs.append(price_diff)
            matched_rows.append({
                'Phone': or_['Phone_Display'],
                'Supplier Date': str(sr.get('Sup_Date','')),
                'Supplier CBD (NIS)': actual_sup,
                'Expected Supplier Price': expected_sup,
                'Price Diff (NIS)': price_diff,
                'Our Tx ID': or_['Transaction ID'],
                'Our Date': or_['Date & Time'],
                'Our Product': or_['Product Name'],
                'Our EUP (NIS)': our_eup,
            })

    for si, sr in sup_matched.iterrows():
        if si in used_sup: continue
        phone = sr['phone_norm']
        actual_sup = sr['CBD']
        om = our_dc[
            (our_dc['phone_norm'] == phone) &
            (~our_dc['Transaction ID'].isin(used_our))
        ]
        if len(om) > 0:
            or_ = om.iloc[0]
            used_our.add(or_['Transaction ID'])
            used_sup.add(si)
            our_eup = or_['End User Price']
            expected_sup = cellcom_expected_supplier_price(our_eup)
            price_diff = round(actual_sup - expected_sup, 2)
            price_diffs.append(price_diff)
            matched_rows.append({
                'Phone': or_['Phone_Display'],
                'Supplier Date': str(sr.get('Sup_Date','')),
                'Supplier CBD (NIS)': actual_sup,
                'Expected Supplier Price': expected_sup,
                'Price Diff (NIS)': price_diff,
                'Our Tx ID': or_['Transaction ID'],
                'Our Date': or_['Date & Time'],
                'Our Product': or_['Product Name'],
                'Our EUP (NIS)': our_eup,
            })

    matched_df = pd.DataFrame(matched_rows)

    if len(matched_df) > 0:
        matched_df['Expected Discount'] = matched_df['Our EUP (NIS)'] - matched_df['Expected Supplier Price']
        total_expected_discount = round(matched_df['Expected Discount'].sum(), 2)
        actual_discount = round(matched_df['Our EUP (NIS)'].sum() - matched_df['Supplier CBD (NIS)'].sum(), 2)
        unexplained_diff = round(actual_discount - total_expected_discount, 2)
        anomaly_rows = matched_df[matched_df['Price Diff (NIS)'].abs() > 0.01].copy()
    else:
        total_expected_discount = 0
        actual_total_diff = 0
        unexplained_diff = 0
        anomaly_rows = pd.DataFrame()

    sup_only_df = sup_pure[sup_pure['phone_norm'].isin(sup_only_phones)].copy()
    sup_only_df['Reason'] = 'Not found in our system'
    sup_only_df['Check_Instruction'] = sup_only_df.apply(
        lambda r: make_check_instruction('Supplier Only', r.get('Sup_Date',''), '', report_date), axis=1)

    our_only_df = our_dc[our_dc['phone_norm'].isin(our_only_phones)].copy()
    our_only_df['Reason'] = our_only_df.apply(
        lambda r: 'Late transaction (after 22:00)' if r.get('Is_Late',False) else 'Not found at supplier', axis=1)
    our_only_df['Check_Instruction'] = our_only_df.apply(
        lambda r: make_check_instruction('Our Only','',r.get('Date & Time',''), report_date, r.get('Is_Late',False)), axis=1)

    t = {
        'sup_cbd': sup_pure[sup_pure['phone_norm'].isin(matched_phones)]['CBD'].sum(),
        'our_eup': our_dc[our_dc['phone_norm'].isin(matched_phones)]['End User Price'].sum(),
        'sup_only_cbd': sup_only_df['CBD'].sum() if len(sup_only_df) else 0,
        'our_only_eup': our_only_df['End User Price'].sum() if len(our_only_df) else 0,
        'refunds_eup': our_refunds['End User Price'].sum() if len(our_refunds) else 0,
        'pending_eup': our_pending['End User Price'].sum() if len(our_pending) else 0,
        'matched_count': len(matched_phones),
        'sup_only_count': len(sup_only_phones),
        'our_only_count': len(our_only_phones),
        'refunds_count': len(our_refunds),
        'unmatched_our_ref_count': len(unmatched_our_ref),
        'failed_count': len(our_failed),
        'pending_count': len(our_pending),
        'price_anomaly_count': len(anomaly_rows),
        'unexplained_diff': unexplained_diff,
        'total_expected_discount': total_expected_discount,
    }
    t['real_gap'] = round(t['sup_only_cbd'] - t['our_only_eup'], 2)

    return {'matched': matched_df, 'sup_only': sup_only_df, 'our_only': our_only_df,
            'refunds': our_refunds, 'failed': our_failed, 'pending': our_pending,
            'anomalies': anomaly_rows, 'totals': t}

# ============================================================
# RECONCILIATION — PELEPHONE
# ============================================================
def run_recon_pelephone(sup_df, pele_df, global_df, esim_df, report_date):
    our_all = pd.concat([pele_df, global_df, esim_df], ignore_index=True)
    our_dc      = our_all[(our_all['Eff_Status'].isin(['DONE','CANCELLED'])) & (~our_all['Is_Refund'])].copy()
    our_refunds = our_all[our_all['Is_Refund']].copy()
    our_failed  = our_all[our_all['Eff_Status'] == 'FAILED'].copy()

    sup_order_ids = set(sup_df['Order_Number'].astype(str).str.strip())
    our_dc['tx_clean'] = our_dc['Transaction ID'].astype(str).str.strip()
    our_tx_ids = set(our_dc['tx_clean'])

    matched_ids  = sup_order_ids & our_tx_ids
    sup_only_ids = sup_order_ids - our_tx_ids
    our_only_ids = our_tx_ids - sup_order_ids

    matched_rows = []
    for _, sr in sup_df[sup_df['Order_Number'].astype(str).str.strip().isin(matched_ids)].iterrows():
        om = our_dc[our_dc['tx_clean'] == sr['Order_Number'].strip()]
        if len(om) > 0:
            or_ = om.iloc[0]
            sup_price = sr['TOPUP_PRICE']
            our_eup   = or_['End User Price']
            diff = round(sup_price - our_eup, 2)
            is_esim = str(or_.get('Operator','')) == 'eSIM'
            matched_rows.append({
                'Phone': or_['Phone_Display'],
                'Supplier Date': f"{sr.get('Sup_Date','')} {sr.get('Sup_Time','')}",
                'Supplier Tx (Doc)': sr.get('Doc_Number',''),
                'Order Number': sr['Order_Number'],
                'Supplier Price (NIS)': sup_price,
                'Our Tx ID': or_['Transaction ID'],
                'Our Date': or_['Date & Time'],
                'Our Operator': or_['Operator'],
                'Our Product': or_['Product Name'],
                'Our EUP (NIS)': our_eup,
                'Difference (NIS)': diff,
                'Note': 'eSIM: expected +2.67 diff' if is_esim and abs(diff - 2.67) < 0.5 else (
                    '⚠️ Unexpected diff' if abs(diff) > 0.01 and not (is_esim and abs(diff-2.67)<0.5) else ''),
            })

    matched_df = pd.DataFrame(matched_rows)
    sup_only_df = sup_df[sup_df['Order_Number'].astype(str).str.strip().isin(sup_only_ids)].copy()
    sup_only_df['Reason'] = 'Order not found in our system'
    sup_only_df['Check_Instruction'] = sup_only_df.apply(
        lambda r: make_check_instruction('Supplier Only',
            f"{r.get('Sup_Date','')} {r.get('Sup_Time','')}", '', report_date), axis=1)

    our_only_df = our_dc[our_dc['tx_clean'].isin(our_only_ids)].copy()
    our_only_df['Reason'] = our_only_df.apply(
        lambda r: 'Late transaction (after 22:00)' if r.get('Is_Late',False) else 'Order not in supplier report', axis=1)
    our_only_df['Check_Instruction'] = our_only_df.apply(
        lambda r: make_check_instruction('Our Only','',r.get('Date & Time',''),report_date,r.get('Is_Late',False)), axis=1)

    esim_diffs = matched_df[matched_df['Note'].str.contains('eSIM', na=False)]
    unexpected = matched_df[matched_df['Note'].str.contains('Unexpected', na=False)]
    esim_rows = (matched_df[matched_df['Our Operator'] == 'eSIM']
                 if 'Our Operator' in matched_df.columns else pd.DataFrame())

    t = {
        'sup_price': matched_df['Supplier Price (NIS)'].sum() if len(matched_df) else 0,
        'our_eup': matched_df['Our EUP (NIS)'].sum() if len(matched_df) else 0,
        'diff': matched_df['Difference (NIS)'].sum() if len(matched_df) else 0,
        'sup_only_price': sup_only_df['TOPUP_PRICE'].sum() if len(sup_only_df) else 0,
        'our_only_eup': our_only_df['End User Price'].sum() if len(our_only_df) else 0,
        'refunds_eup': our_refunds['End User Price'].sum() if len(our_refunds) else 0,
        'matched_count': len(matched_ids),
        'sup_only_count': len(sup_only_ids),
        'our_only_count': len(our_only_ids),
        'refunds_count': len(our_refunds),
        'failed_count': len(our_failed),
        'esim_diff_count': len(esim_diffs),
        'esim_diff_total': round(esim_diffs['Difference (NIS)'].sum(), 2) if len(esim_diffs) else 0,
        'esim_count': len(esim_rows),
        'esim_our_total': round(esim_rows['Our EUP (NIS)'].sum(), 2) if len(esim_rows) else 0,
        'esim_sup_total': round(esim_rows['Supplier Price (NIS)'].sum(), 2) if len(esim_rows) else 0,
        'unexpected_diff_count': len(unexpected),
    }
    t['real_gap'] = round(t['sup_only_price'] - t['our_only_eup'], 2)

    return {'matched': matched_df, 'sup_only': sup_only_df, 'our_only': our_only_df,
            'refunds': our_refunds, 'failed': our_failed, 'totals': t}

# ============================================================
# RAVKAV
# ============================================================
RK_CHARGED = 'חויב'
RK_FAILED = 'נכשל'
RK_REFUNDED = 'זוכה'

def load_ravkav_our(file_bytes):
    """Our System Logs export for RavKav. Extracts the RavKav uid and card
    serial from the embedded JSON so transactions match 1:1 with RavKav."""
    try:
        text = None
        for enc in ['utf-8-sig', 'utf-8', 'windows-1255', 'cp1255', 'latin1']:
            try:
                text = file_bytes.decode(enc); break
            except Exception:
                continue
        df = pd.read_csv(StringIO(text), dtype=str, on_bad_lines='skip')
        df['End User Price'] = pd.to_numeric(df.get('End User Price', 0), errors='coerce').fillna(0)
        if 'Error details' in df.columns:
            det = df['Error details'].fillna('')
        else:
            det = pd.Series([''] * len(df), index=df.index)
        df['uid'] = det.str.extract(r'"uid": "([0-9a-fA-F\-]{36})"', expand=False).fillna('')
        df['card_serial'] = det.str.extract(r'"card_serial": "([^"]+)"', expand=False).fillna('')
        try:
            df['_dt'] = pd.to_datetime(df['Date & Time'], dayfirst=True, errors='coerce')
        except Exception:
            df['_dt'] = pd.NaT
        return df, None
    except Exception as e:
        return None, str(e)

def load_ravkav_totals(file_bytes):
    """RavKav daily one-line totals report. Amounts arrive in agorot -> NIS."""
    try:
        text = None
        for enc in ['utf-8-sig', 'utf-8', 'windows-1255', 'cp1255', 'latin1']:
            try:
                text = file_bytes.decode(enc); break
            except Exception:
                continue
        df = pd.read_csv(StringIO(text))
        if len(df) == 0:
            return None, "Totals file is empty"
        needed = ['Total Charged Amount', 'Total Refunded Amount', 'Net Total Charged Amount']
        missing = [c for c in needed if c not in df.columns]
        if missing:
            return None, f"Missing columns: {missing}"
        row = df.iloc[0]
        def agorot(col):
            try:
                return round(float(row.get(col, 0) or 0) / 100.0, 2)
            except Exception:
                return 0.0
        return {
            'charged': agorot('Total Charged Amount'),
            'refunded': agorot('Total Refunded Amount'),
            'net': agorot('Net Total Charged Amount'),
            'balance_charged': agorot('Balance Charged Amount'),
            'cc_charged': agorot('Credit Card Charged Amount'),
        }, None
    except Exception as e:
        return None, str(e)

def load_ravkav_detailed(file_bytes):
    """RavKav detailed transaction report (xlsx) — requested on discrepancies."""
    try:
        df = pd.read_excel(BytesIO(file_bytes))
        needed = ['uid', 'transaction_status', 'charged_amount', 'card__serial_number']
        missing = [c for c in needed if c not in df.columns]
        if missing:
            return None, f"Missing columns: {missing}"
        df['uid'] = df['uid'].astype(str).str.strip()
        df['charged_amount'] = pd.to_numeric(df['charged_amount'], errors='coerce').fillna(0)
        df['card__serial_number'] = df['card__serial_number'].astype(str)
        return df, None
    except Exception as e:
        return None, str(e)

def run_recon_ravkav(our_df, sup_totals, sup_det, report_date):
    """Totals comparison + optional uid-level detailed comparison."""
    act = our_df['Action'] if 'Action' in our_df.columns else 'PURCHASE'
    stat = our_df['Status'] if 'Status' in our_df.columns else ''
    done_all = our_df[(act == 'PURCHASE') & (stat == 'DONE')].copy()
    done = done_all[done_all['uid'] != ''].copy()          # comparable with RavKav
    adjustments = done_all[done_all['uid'] == ''].copy()   # anomaly: DONE purchase without uid
    failed = our_df[stat == 'FAILED'].copy()
    cancelled = our_df[stat == 'CANCELLED'].copy()         # refunded on RavKav side (זוכה)
    client_refunds = our_df[(act == 'REFUND') & (stat == 'DONE')].copy()  # our credits to clients, no uid

    our_charged = round(float(done['End User Price'].sum()), 2)
    our_refunded = round(float(cancelled['End User Price'].sum()), 2)
    client_ref_sum = round(float(client_refunds['End User Price'].sum()), 2) if len(client_refunds) else 0.0

    t = {
        'done_count': len(done), 'failed_count': len(failed),
        'cancelled_count': len(cancelled),
        'adj_count': len(adjustments),
        'adj_sum': round(float(adjustments['End User Price'].sum()), 2) if len(adjustments) else 0.0,
        'our_charged': our_charged, 'our_refunded': our_refunded,
        'sup_charged': None, 'sup_refunded': None, 'sup_net': None,
        'gap_charged': 0.0, 'gap_refunded': 0.0, 'gap_net': 0.0,
        'client_ref_count': len(client_refunds), 'client_ref_sum': client_ref_sum,
        'matched_count': len(done), 'pending_count': 0,
        'refunds_count': len(client_refunds), 'refunds_eup': client_ref_sum,
    }
    if sup_totals:
        t['sup_charged'] = sup_totals['charged']
        t['sup_refunded'] = sup_totals['refunded']
        t['sup_net'] = sup_totals['net']
        t['gap_charged'] = round(sup_totals['charged'] - our_charged, 2)
        t['gap_refunded'] = round(sup_totals['refunded'] - our_refunded, 2)
        t['gap_net'] = round(sup_totals['net'] - (our_charged - our_refunded), 2)

    detailed = None
    sup_rows, our_rows = [], []
    if sup_det is not None:
        ours_uid = {r['uid']: r for _, r in our_df[our_df['uid'] != ''].iterrows()}
        theirs_uid = {r['uid']: r for _, r in sup_det.iterrows()}
        both = set(ours_uid) & set(theirs_uid)
        only_ours_ids = sorted(set(ours_uid) - set(theirs_uid))
        only_theirs_ids = sorted(set(theirs_uid) - set(ours_uid))

        failed_but_charged, done_not_charged, amount_mismatch = [], [], []
        agree = 0
        for u in both:
            o, s = ours_uid[u], theirs_uid[u]
            s_charged = s['transaction_status'] == RK_CHARGED
            if o['Status'] == 'FAILED' and s_charged:
                failed_but_charged.append((o, s))
            elif o['Status'] == 'DONE' and s['transaction_status'] == RK_FAILED:
                done_not_charged.append((o, s))
            elif o['Status'] == 'DONE' and s_charged \
                    and abs(float(o['End User Price']) - float(s['charged_amount'])) > 0.01:
                amount_mismatch.append((o, s))
            else:
                agree += 1

        def _disp(pairs):
            return pd.DataFrame([{
                'uid': str(o.get('uid') or s.get('uid'))[:13] + '…',
                'Card': str(s.get('card__serial_number', o.get('card_serial', ''))),
                'Our status': o.get('Status', ''),
                'RavKav status': s.get('transaction_status', ''),
                'Our client price (NIS)': float(o.get('End User Price', 0)),
                'RavKav amount (NIS)': float(s.get('charged_amount', 0)),
                'Date & Time': str(o.get('Date & Time', '')),
            } for o, s in pairs])

        detailed = {
            'matched_count': agree,
            'failed_but_charged': _disp(failed_but_charged),
            'done_not_charged': _disp(done_not_charged),
            'amount_mismatch': _disp(amount_mismatch),
            'only_ours': pd.DataFrame([{
                'uid': u[:13] + '…',
                'Card': str(ours_uid[u].get('card_serial', '')),
                'Our status': ours_uid[u].get('Status', ''),
                'Our client price (NIS)': float(ours_uid[u].get('End User Price', 0)),
                'Date & Time': str(ours_uid[u].get('Date & Time', '')),
            } for u in only_ours_ids]),
            'only_theirs': pd.DataFrame([{
                'uid': u[:13] + '…',
                'Card': str(theirs_uid[u].get('card__serial_number', '')),
                'RavKav status': theirs_uid[u].get('transaction_status', ''),
                'RavKav amount (NIS)': float(theirs_uid[u].get('charged_amount', 0)),
            } for u in only_theirs_ids]),
            'adjustments': (adjustments[['Date & Time', 'Transaction ID', 'End User Price']]
                            .rename(columns={'End User Price': 'Client price (NIS)'})
                            if len(adjustments) else pd.DataFrame()),
        }

        for o, s in failed_but_charged:
            sup_rows.append({
                'Phone_Display': str(s.get('card__serial_number', '')),
                'Sup_Date': str(o.get('Date & Time', report_date)), 'Package': 'RavKav',
                'CBD': float(s.get('charged_amount', 0)),
                'Reason': 'Charged by RavKav, FAILED in our system',
                'Check_Instruction': 'Verify the card was actually reloaded; '
                                     'if yes — fix our status, if no — request a refund from RavKav.',
            })
        for u in only_theirs_ids:
            s = theirs_uid[u]
            sup_rows.append({
                'Phone_Display': str(s.get('card__serial_number', '')),
                'Sup_Date': str(report_date), 'Package': 'RavKav',
                'CBD': float(s.get('charged_amount', 0)),
                'Reason': 'At RavKav only — missing in our log',
                'Check_Instruction': 'Transaction exists at RavKav but not in our system — investigate with RavKav.',
            })
        for o, s in done_not_charged:
            our_rows.append({
                'Phone_Display': str(o.get('Phone Number', '')), 'Operator': 'RavKav',
                'Product Name': 'RavKav', 'End User Price': float(o.get('End User Price', 0)),
                'Date & Time': str(o.get('Date & Time', '')),
                'Reason': 'DONE in our system, failed at RavKav',
                'Check_Instruction': 'Verify whether the client was charged; align statuses with RavKav.',
            })
        for u in only_ours_ids:
            o = ours_uid[u]
            our_rows.append({
                'Phone_Display': str(o.get('Phone Number', '')), 'Operator': 'RavKav',
                'Product Name': 'RavKav', 'End User Price': float(o.get('End User Price', 0)),
                'Date & Time': str(o.get('Date & Time', '')),
                'Reason': 'In our log only — uid unknown to RavKav',
                'Check_Instruction': 'Verify the transaction with RavKav.',
            })
        for _, o in adjustments.iterrows():
            our_rows.append({
                'Phone_Display': str(o.get('Phone Number', '')), 'Operator': 'RavKav',
                'Product Name': 'Manual adjustment', 'End User Price': float(o.get('End User Price', 0)),
                'Date & Time': str(o.get('Date & Time', '')),
                'Reason': 'Manual adjustment (no RavKav uid)',
                'Check_Instruction': 'Confirm the adjustment is intentional and documented.',
            })

    sup_only_df = pd.DataFrame(sup_rows)
    our_only_df = pd.DataFrame(our_rows)
    t['sup_only_count'] = len(sup_only_df)
    t['sup_only_cbd'] = round(float(sup_only_df['CBD'].sum()), 2) if len(sup_only_df) else 0.0
    t['our_only_count'] = len(our_only_df)
    t['our_only_eup'] = round(float(our_only_df['End User Price'].sum()), 2) if len(our_only_df) else 0.0
    if sup_totals:
        t['real_gap'] = t['gap_net']
    elif sup_det is not None:
        their_charged = round(float(
            sup_det[sup_det['transaction_status'] == RK_CHARGED]['charged_amount'].sum()), 2)
        t['real_gap'] = round(their_charged - our_charged, 2)
    else:
        t['real_gap'] = 0.0

    refunds_df = pd.DataFrame(
        [{
            'Operator': 'RavKav', 'Phone_Display': str(r.get('Phone Number', '')),
            'Date & Time': str(r.get('Date & Time', '')),
            'Product Name': 'Client refund (internal)',
            'End User Price': float(r.get('End User Price', 0)),
        } for _, r in client_refunds.iterrows()] +
        [{
            'Operator': 'RavKav', 'Phone_Display': str(r.get('Phone Number', '')),
            'Date & Time': str(r.get('Date & Time', '')),
            'Product Name': 'Refunded by RavKav',
            'End User Price': float(r.get('End User Price', 0)),
        } for _, r in cancelled.iterrows()])

    return {'matched': pd.DataFrame(), 'sup_only': sup_only_df, 'our_only': our_only_df,
            'refunds': refunds_df, 'failed': failed, 'totals': t, 'detailed': detailed}

# ============================================================
# SHARED UI COMPONENTS
# ============================================================
def render_header(title, subtitle, logos, extra_labels=None):
    logo_html = ''.join([
        '<img src="' + l + '" style="height:48px;object-fit:contain;">'
        for l in logos if l
    ])
    if extra_labels:
        for label in extra_labels:
            logo_html += '<span style="color:white;font-size:20px;font-weight:600;letter-spacing:1px;padding:4px 10px;border:2px solid rgba(255,255,255,0.6);border-radius:6px;">' + label + '</span>'
    st.markdown(f"""
    <div class="main-header">
        <div style="flex:1">
            <div style="display:flex;align-items:center;gap:12px;margin-bottom:6px;">
                {logo_html}
            </div>
            <h1>{title}</h1>
            <p>{subtitle}</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

def render_action_required(sup_only_df, our_only_df, report_date, shifts_our=None, shifts_sup=None):
    sup_count = len(sup_only_df)
    our_count = len(our_only_df)
    total = sup_count + our_count
    if total == 0:
        st.success("✅ No verification needed — all transactions matched!")
        return

    st.markdown(f"""
    <div class="action-box">
        <h4>⚠️ Action Required: {total} phone(s) need verification</h4>
    </div>
    """, unsafe_allow_html=True)

    rows = []
    if len(sup_only_df) > 0:
        for _, r in sup_only_df.iterrows():
            phone = r.get('Phone_Display', r.get('phone_norm',''))
            date_val = str(r.get('Sup_Date', r.get('sup_date','')))
            product = r.get('Package', r.get('TOPUP_ITEM', r.get('Product Name','')))
            amount = r.get('CBD', r.get('TOPUP_PRICE', 0))
            instruction = r.get('Check_Instruction','Check our reports for this date')
            note = ''
            if shifts_sup and phone in (shifts_sup or []):
                note = '✅ Date shift confirmed (found in yesterday our-only)'
            rows.append({
                'Phone': phone, 'Date': date_val, 'Product': product,
                'Amount (NIS)': amount, 'Where': '❌ In supplier — not in ours',
                'What to check': instruction, 'Auto-check': note,
                'Verified': '⬜ Not checked'
            })
    if len(our_only_df) > 0:
        for _, r in our_only_df.iterrows():
            phone = r.get('Phone_Display','')
            date_val = r.get('Date & Time','')
            product = r.get('Product Name','')
            amount = r.get('End User Price',0)
            instruction = r.get('Check_Instruction','Check supplier report for next day')
            note = ''
            if shifts_our and phone in shifts_our:
                note = '✅ Date shift confirmed (found in yesterday supplier-only)'
            rows.append({
                'Phone': phone, 'Date': date_val, 'Product': product,
                'Amount (NIS)': amount, 'Where': '⚠️ In ours — not in supplier',
                'What to check': instruction, 'Auto-check': note,
                'Verified': '⬜ Not checked'
            })
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

def render_summary_tab(sup_only_df, our_only_df, t, report_date, tab_name):
    st.markdown("### 📋 Discrepancy Summary")
    col1, col2, col3 = st.columns(3)
    col1.metric("❌ Supplier Only", t['sup_only_count'],
                delta=f"{t['sup_only_cbd'] if 'sup_only_cbd' in t else t.get('sup_only_price',0):,.2f} NIS",
                delta_color="inverse")
    col2.metric("⚠️ Our Only", t['our_only_count'],
                delta=f"{t['our_only_eup']:,.2f} NIS", delta_color="inverse")
    gap = t.get('real_gap',0)
    col3.metric("📊 Real Gap", f"{gap:,.2f} NIS",
                delta="Sup Only − Our Only",
                delta_color="inverse" if gap > 0 else "normal")

    st.markdown("---")
    combined = []
    if len(sup_only_df) > 0:
        for _, r in sup_only_df.iterrows():
            combined.append({
                'Side': '❌ Supplier Only',
                'Date': str(r.get('Sup_Date', r.get('sup_date',''))),
                'Phone': r.get('Phone_Display', r.get('phone_norm','')),
                'Product': r.get('Package', r.get('TOPUP_ITEM','')),
                'Amount (NIS)': r.get('CBD', r.get('TOPUP_PRICE',0)),
                'What to check': r.get('Check_Instruction',''),
            })
    if len(our_only_df) > 0:
        for _, r in our_only_df.iterrows():
            combined.append({
                'Side': '⚠️ Our Only',
                'Date': r.get('Date & Time',''),
                'Phone': r.get('Phone_Display',''),
                'Product': r.get('Product Name',''),
                'Amount (NIS)': r.get('End User Price',0),
                'What to check': r.get('Check_Instruction',''),
            })
    if combined:
        st.dataframe(pd.DataFrame(combined), use_container_width=True, hide_index=True)
    else:
        st.success("✅ No discrepancies!")

def build_detail_rows(report_date, operator_tab, sup_only_df, our_only_df):
    rows = []
    if len(sup_only_df) > 0:
        for _, r in sup_only_df.iterrows():
            rows.append([
                report_date, operator_tab, 'Supplier Only',
                r.get('Phone_Display', r.get('phone_norm','')),
                '', r.get('Package', r.get('TOPUP_ITEM','')),
                r.get('CBD', r.get('TOPUP_PRICE',0)),
                str(r.get('Sup_Date','')), '',
                r.get('Reason',''), r.get('Check_Instruction',''), '⬜ Not checked',
                '', ''
            ])
    if len(our_only_df) > 0:
        for _, r in our_only_df.iterrows():
            rows.append([
                report_date, operator_tab, 'Our Only',
                r.get('Phone_Display',''),
                r.get('Operator',''), r.get('Product Name',''),
                r.get('End User Price',0),
                '', r.get('Date & Time',''),
                r.get('Reason',''), r.get('Check_Instruction',''), '⬜ Not checked',
                '', ''
            ])
    return rows

# ============================================================
# MAIN APP
# ============================================================
USERS_SHEET = 'Users'
USERS_COLS = ['email', 'pass_hash', 'role', 'active', 'created_by', 'created_at']

def _hash_pw(p):
    import hashlib
    return hashlib.sha256(str(p).encode()).hexdigest()

def _now_il():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo('Asia/Jerusalem')).strftime('%Y-%m-%d %H:%M')
    except Exception:
        return datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')

def _users_ws():
    sh = get_spreadsheet('partner')
    if sh is None:
        return None
    return get_or_create_sheet(sh, USERS_SHEET, USERS_COLS)

@st.cache_data(ttl=60, show_spinner=False)
def load_users():
    """All user rows with their sheet row number (_row). Cached 60s —
    deactivation takes effect within a minute."""
    ws = _users_ws()
    if ws is None:
        return []
    vals = _api_retry(lambda: ws.get_all_values())
    if not vals:
        return []
    hdr = vals[0]
    out = []
    for i, rv in enumerate(vals[1:], start=2):
        r = dict(zip(hdr, rv))
        r['_row'] = i
        out.append(r)
    return out

def _is_active(u):
    return str(u.get('active', '')).strip().lower() in ('1', 'true', 'yes', 'y')

def _send_daily_email(op_label, report_date, result, t):
    """Email the daily Excel report to CS. Configured via secrets [email].
    Never breaks the save flow: info note if not configured, warning on failure."""
    try:
        cfg = dict(st.secrets["email"])
    except Exception:
        st.info("📧 Email not configured — report not sent (add [email] to secrets).")
        return
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.application import MIMEApplication

        def _clean(s):
            # removes ALL whitespace, including non-breaking spaces (\xa0)
            # that Google inserts between App Password groups
            return ''.join(str(s).split())
        user = _clean(cfg.get('smtp_user', ''))
        pw = _clean(cfg.get('smtp_app_password', ''))
        rcpts = [_clean(r) for r in
                 str(cfg.get('recipients', '')).replace(';', ',').split(',')
                 if _clean(r)]
        if not user or not pw or not rcpts:
            st.warning("📧 Email config incomplete (need smtp_user, smtp_app_password, recipients).")
            return

        try:
            _d = datetime.strptime(str(report_date), '%Y-%m-%d').strftime('%d.%m.%Y')
        except Exception:
            _d = str(report_date)
        subj = f"{op_label} - Daily Reconciliation {_d}"
        gap = t.get('real_gap', 0)
        body = (
            f"Daily reconciliation — {op_label} — {report_date}\n\n"
            f"Matched: {t.get('matched_count', 0)}\n"
            f"Supplier Only: {t.get('sup_only_count', 0)} "
            f"({t.get('sup_only_cbd', t.get('sup_only_price', 0)):,.2f} NIS)\n"
            f"Our Only: {t.get('our_only_count', 0)} ({t.get('our_only_eup', 0):,.2f} NIS)\n"
            f"Real Gap: {gap:,.2f} NIS\n"
            f"Refunds: {t.get('refunds_count', 0)} ({t.get('refunds_eup', 0):,.2f} NIS)\n"
            f"Pending: {t.get('pending_count', 0)}\n\n"
            f"Saved by: {_current_user()} at {_now_il()}\n"
            f"Full Excel report attached."
        )
        if t.get('esim_count'):
            _ed = t.get('esim_sup_total', 0) - t.get('esim_our_total', 0)
            body += (f"\n\neSIM: {t.get('esim_count')} transactions | "
                     f"our {t.get('esim_our_total', 0):,.2f} NIS | "
                     f"supplier {t.get('esim_sup_total', 0):,.2f} NIS | "
                     f"price diff {_ed:,.2f} NIS")
        msg = MIMEMultipart()
        msg['From'] = user
        msg['To'] = ', '.join(rcpts)
        msg['Subject'] = subj
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        buf = create_excel_report(result, report_date, op_label)
        att = MIMEApplication(buf.getvalue())
        fname = (op_label.replace(' ', '_').replace('&', 'and')
                 + '_' + str(report_date).replace('-', '_') + '.xlsx')
        att.add_header('Content-Disposition', 'attachment', filename=fname)
        msg.attach(att)

        with smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=30) as srv:
            srv.login(user, pw)
            srv.sendmail(user, rcpts, msg.as_string())
        st.success(f"📧 Report emailed to {', '.join(rcpts)}")
    except Exception as e:
        st.warning(f"📧 Email failed (data was saved OK): {e}")

SESSION_HOURS = 6

def _session_key():
    try:
        return str(st.secrets["gcp_service_account"]["private_key"])
    except Exception:
        return "local-dev-fallback-key"

def _make_session_token(email, src, hours=SESSION_HOURS):
    import hmac, hashlib, base64, time
    exp = int(time.time()) + int(hours * 3600)
    payload = f"{email}|{src}|{exp}"
    sig = hmac.new(_session_key().encode(), payload.encode(),
                   hashlib.sha256).hexdigest()[:32]
    return base64.urlsafe_b64encode(f"{payload}|{sig}".encode()).decode()

def _parse_session_token(tok):
    """Returns (email, src) if the token is authentic and not expired, else None."""
    import hmac, hashlib, base64, time
    try:
        raw = base64.urlsafe_b64decode(str(tok).encode()).decode()
        parts = raw.split('|')
        if len(parts) != 4:
            return None
        email, src, exp, sig = parts
        payload = f"{email}|{src}|{exp}"
        good = hmac.new(_session_key().encode(), payload.encode(),
                        hashlib.sha256).hexdigest()[:32]
        if not hmac.compare_digest(sig, good):
            return None
        if int(exp) < time.time():
            return None
        return email.strip().lower(), src
    except Exception:
        return None

def check_login():
    """Returns username (email) if logged in, 'setup' if auth is not
    configured yet (open access, setup mode), or None while the
    login form is displayed."""
    try:
        fb = dict(st.secrets["auth_admin"])
    except Exception:
        fb = {}
    users = load_users()
    active_users = [u for u in users if _is_active(u)]

    if not fb and not active_users:
        return "setup"   # nothing configured -> open, create first admin in UI

    # Page refresh wipes session_state -> try to restore from the signed token
    if not st.session_state.get('auth_user'):
        _tok = st.query_params.get('s')
        if _tok:
            _te = _parse_session_token(_tok)
            if _te:
                _e, _src = _te
                if _src == 'secrets' and fb \
                        and _e == str(fb.get('email', '')).strip().lower():
                    st.session_state['auth_user'] = _e
                    st.session_state['auth_role'] = 'owner'
                    st.session_state['auth_src'] = 'secrets'
                else:
                    _m0 = [u for u in active_users
                           if str(u.get('email', '')).strip().lower() == _e]
                    if _m0:
                        st.session_state['auth_user'] = _e
                        st.session_state['auth_role'] = (_m0[0].get('role') or 'user').strip().lower()
                        st.session_state['auth_src'] = 'sheet'
            if not st.session_state.get('auth_user'):
                try:
                    del st.query_params['s']   # invalid, expired or deactivated
                except Exception:
                    pass

    cur = st.session_state.get('auth_user')
    if cur:
        if st.session_state.get('auth_src') == 'sheet':
            m = [u for u in active_users
                 if str(u.get('email', '')).strip().lower() == cur]
            if not m:
                st.session_state.pop('auth_user', None)
                st.session_state.pop('auth_role', None)
                st.session_state.pop('auth_src', None)
                st.warning("Your access was deactivated. Contact an administrator.")
            else:
                st.session_state['auth_role'] = (m[0].get('role') or 'user').strip().lower()
                return cur
        else:
            return cur

    st.markdown("## 🔐 Reconciliation Dashboard — Sign in")
    with st.form("login_form"):
        em = st.text_input("Email")
        pw = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Sign in", type="primary")
    if submitted:
        e = em.strip().lower()
        h = _hash_pw(pw)
        if fb and e == str(fb.get('email', '')).strip().lower() \
              and h == str(fb.get('pass_hash', '')).strip().lower():
            st.session_state['auth_user'] = e
            st.session_state['auth_role'] = 'owner'
            st.session_state['auth_src'] = 'secrets'
            st.query_params['s'] = _make_session_token(e, 'secrets')
            st.rerun()
        m = [u for u in active_users
             if str(u.get('email', '')).strip().lower() == e
             and str(u.get('pass_hash', '')).strip().lower() == h]
        if m:
            st.session_state['auth_user'] = e
            st.session_state['auth_role'] = (m[0].get('role') or 'user').strip().lower()
            st.session_state['auth_src'] = 'sheet'
            st.query_params['s'] = _make_session_token(e, 'sheet')
            st.rerun()
        else:
            st.error("Invalid email or password")
    return None

def main():
    _user = check_login()
    if _user is None:
        return
    if _user == "setup":
        st.session_state['auth_role'] = 'owner'
        st.session_state['auth_setup_mode'] = True
    else:
        st.session_state.pop('auth_setup_mode', None)

    with st.sidebar:
        st.markdown("### 📋 Navigation")
        if st.session_state.get('auth_setup_mode'):
            st.warning("⚠️ Open access — create the first admin in 👥 Users")
        if st.session_state.get('auth_user'):
            _uc1, _uc2 = st.columns([3, 2])
            _uc1.caption(f"👤 {st.session_state['auth_user']} ({st.session_state.get('auth_role', 'user')})")
            if _uc2.button("Logout", key="logout_btn"):
                for _k in ('auth_user', 'auth_role', 'auth_src'):
                    st.session_state.pop(_k, None)
                try:
                    del st.query_params['s']
                except Exception:
                    pass
                st.rerun()
        if LOGO_PAYX:
            st.markdown(f'<img src="{LOGO_PAYX}" style="height:28px;margin-bottom:8px;">', unsafe_allow_html=True)
        _pages = [
            "📱 Partner + 012Talk Reconciliation",
            "⭐ Pelephone Reconciliation",
            "📡 Cellcom Reconciliation",
            "🚌 RavKav Reconciliation",
            "📅 Monthly Summary",
            "⏳ Pending Verification",
            "✅ Verified",
            "ℹ️ Instructions",
        ]
        if st.session_state.get('auth_role') in ('admin', 'owner'):
            _pages.append("👥 Users")
        page = st.radio("Select page", _pages, label_visibility="collapsed")

        st.markdown("---")
        st.markdown("### 📊 History")

        # Cached connectivity check — only once per session
        if 'sheets_ok' not in st.session_state:
            st.session_state['sheets_ok'] = get_spreadsheet('partner') is not None
        if st.session_state['sheets_ok']:
            st.success("✅ Google Sheets connected")
        else:
            st.error("⛔ Google Sheets unavailable\nData will NOT be saved to the cloud!")

        # Cache history stats — only load once per session
        if 'sidebar_stats' not in st.session_state:
            total_days = 0
            last_date = None
            for op in ['partner', 'pelephone', 'cellcom', 'ravkav']:
                h = load_history(operator_tab=op)
                total_days += len(h)
                if h and (last_date is None or h[-1].get('date','') > last_date):
                    last_date = h[-1].get('date','')
            st.session_state['sidebar_stats'] = (total_days, last_date)
        total_days, last_date = st.session_state['sidebar_stats']
        if total_days > 0:
            st.success(f"✅ {total_days} days recorded")
            if last_date:
                st.info(f"Last: {last_date}")
        else:
            st.info("No history yet")

        st.markdown("---")
        _build = None
        try:
            import subprocess as _sp
            _out = _sp.getoutput('git rev-parse --short HEAD 2>/dev/null').strip()
            if _out and ' ' not in _out and len(_out) <= 12:
                _build = _out
        except Exception:
            pass
        st.caption(f"Build: {_build or 'n/a'} | {APP_VERSION}")

    # ============================================================
    # PAGE: PARTNER + 012TALK
    # ============================================================
    # Clear save flags when switching pages
    for _key in ['pt_do_save', 'pe_do_save', 'ce_do_save', 'rk_do_save']:
        if _key in st.session_state and not page.startswith(
            {'pt': '📱', 'pe': '⭐', 'ce': '📡', 'rk': '🚌'}[_key[:2]]):
            st.session_state.pop(_key, None)

    if page == "📱 Partner + 012Talk Reconciliation":
        render_header(
            "Partner + 012Talk Reconciliation",
            "Supplier vs Our System — Partner + 012Talk",
            [LOGO_PAYX, LOGO_012],
            extra_labels=['+Partner']
        )
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("**1️⃣ Supplier File (.xls)**")
            sup_file = st.file_uploader("Supplier", type=['xls','xlsx','csv'],
                                        label_visibility="collapsed", key="pt_sup")
            st.caption("Supplier report — Charge Before Discount (col M)")
        with col2:
            st.markdown("**2️⃣ Partner EPRS File (.csv)**")
            part_file = st.file_uploader("Partner EPRS", type=['csv','xlsx'],
                                         label_visibility="collapsed", key="pt_part")
            st.caption("Our system export — Partner operator")
        with col3:
            st.markdown("**3️⃣ 012Talk EPRS File (.csv)**")
            talk_file = st.file_uploader("012Talk EPRS", type=['csv','xlsx'],
                                         label_visibility="collapsed", key="pt_talk")
            st.caption("Our system export — 012Talk operator")

        if sup_file and part_file and talk_file:
            if st.button("▶ Run Reconciliation", type="primary", use_container_width=True, key="pt_run"):
                with st.spinner("Processing..."):
                    sup_df,  e1 = load_supplier_partner(sup_file.read(), sup_file.name)
                    part_df, e2 = load_our(part_file.read(), 'Partner')
                    talk_df, e3 = load_our(talk_file.read(), '012Talk')
                    if e1: st.error(f"Supplier error: {e1}"); return
                    if e2: st.error(f"Partner error: {e2}"); return
                    if e3: st.error(f"012Talk error: {e3}"); return

                    auto_date = date.today().strftime('%Y-%m-%d')
                    if 'Sup_Date' in sup_df.columns and len(sup_df) > 0:
                        try:
                            auto_date = pd.to_datetime(sup_df['Sup_Date'].iloc[0], dayfirst=True).strftime('%Y-%m-%d')
                            st.info(f"📅 Date detected: **{auto_date}**")
                        except: pass

                    result = run_recon_partner(sup_df, part_df, talk_df, auto_date)
                    shifts_our, shifts_sup = cross_day_match(result, auto_date, 'partner')
                    st.session_state['pt_result'] = result
                    st.session_state['pt_date'] = auto_date
                    st.session_state['pt_shifts_our'] = shifts_our
                    st.session_state['pt_shifts_sup'] = shifts_sup
                    st.success("✅ Complete!")

        if 'pt_result' in st.session_state:
            result = st.session_state['pt_result']
            t = result['totals']
            rdate = st.session_state['pt_date']
            shifts_our = st.session_state.get('pt_shifts_our',[])
            shifts_sup = st.session_state.get('pt_shifts_sup',[])

            st.markdown("---")
            c1,c2,c3,c4,c5 = st.columns(5)
            c1.metric("✅ Matched", f"{t['matched_count']:,}")
            c2.metric("❌ Supplier Only", t['sup_only_count'],
                      delta=f"{t['sup_only_cbd']:,.2f} NIS", delta_color="inverse")
            c3.metric("⚠️ Our Only", t['our_only_count'],
                      delta=f"{t['our_only_eup']:,.2f} NIS", delta_color="inverse")
            c4.metric("🔄 Sup vs Pending", t['sup_pnd_count'],
                      delta=f"{t['sup_pnd_cbd']:,.2f} NIS", delta_color="off")
            gap = t['real_gap']
            c5.metric("📊 Real Gap",
                      f"+{gap:,.2f} NIS (sup higher)" if gap>0 else (f"{gap:,.2f} NIS (we higher)" if gap<0 else "0.00 NIS ✅"),
                      delta_color="inverse" if gap>0 else "normal")

            if shifts_our:
                st.success(f"✅ Date Shift Confirmed: {len(shifts_our)} phone(s) from Our Only found in yesterday's Supplier Only — {', '.join(shifts_our)}")
            if shifts_sup:
                st.success(f"✅ Date Shift Confirmed: {len(shifts_sup)} phone(s) from Supplier Only found in yesterday's Our Only — {', '.join(shifts_sup)}")

            render_action_required(result['sup_only'], result['our_only'], rdate, shifts_our, shifts_sup)

            st.markdown("---")
            tabs = st.tabs([
                "📋 Summary",
                f"❌ Supplier Only ({t['sup_only_count']})",
                f"⚠️ Our Only ({t['our_only_count']})",
                f"🔄 Sup vs Pending ({t['sup_pnd_count']})",
                f"✅ Matched ({t['matched_count']})",
                f"🕐 Pending ({t['pending_count']})",
                f"↩️ Refunds ({t['refunds_count']})",
                f"🔴 Failed ({t['failed_count']})",
            ])
            with tabs[0]:
                render_summary_tab(result['sup_only'], result['our_only'], t, rdate, 'partner')
            with tabs[1]:
                if len(result['sup_only']) > 0:
                    _so_cols = [c for c in ['Phone_Display','Sup_Date','Package','CBD','Check_Instruction'] if c in result['sup_only'].columns]
                    show = result['sup_only'][_so_cols].copy()
                    show.rename(columns={'Phone_Display':'Phone','Sup_Date':'Date','CBD':'CBD (NIS)','Check_Instruction':'What to check'}, inplace=True)
                    st.dataframe(show, use_container_width=True, hide_index=True)
                    st.info(f"Total: {t['sup_only_count']} phones | {t['sup_only_cbd']:,.2f} NIS")
                else: st.success("✅ No supplier-only records!")
            with tabs[2]:
                if len(result['our_only']) > 0:
                    late = result['our_only'].get('Is_Late', pd.Series(dtype=bool)).sum() if 'Is_Late' in result['our_only'].columns else 0
                    if late > 0: st.warning(f"⏰ {late} transaction(s) after 22:00 — likely in tomorrow's supplier report")
                    show = result['our_only'][['Phone_Display','Date & Time','Operator','Product Name','End User Price','Check_Instruction']].copy()
                    show.columns = ['Phone','Date & Time','Operator','Product','Client price (NIS)','What to check']
                    st.dataframe(show, use_container_width=True, hide_index=True)
                    st.info(f"Total: {t['our_only_count']} phones | {t['our_only_eup']:,.2f} NIS")
                else: st.success("✅ No our-only records!")
            with tabs[3]:
                if len(result['sup_pending']) > 0:
                    st.dataframe(result['sup_pending'][['Phone_Display','Sup_Date','Package','CBD']].rename(
                        columns={'Phone_Display':'Phone','Sup_Date':'Date','CBD':'CBD (NIS)'}),
                        use_container_width=True, hide_index=True)
                else: st.success("✅ No supplier vs pending conflicts!")
            with tabs[4]:
                if len(result['matched']) > 0:
                    st.dataframe(result['matched'].rename(
                        columns={'Our EUP (NIS)': 'Our Client Price (NIS)'}),
                        use_container_width=True, hide_index=True)
                else: st.info("No matched records")
            with tabs[5]:
                if len(result['pending']) > 0:
                    st.warning(f"⚠️ {t['pending_count']} pending")
                    show = result['pending'][['Phone_Display','Date & Time','Operator','Product Name','End User Price']].rename(columns={'Phone_Display':'Phone','End User Price':'Client price (NIS)'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.success("✅ No pending!")
            with tabs[6]:
                if len(result['refunds']) > 0:
                    show = result['refunds'][['Operator','Phone_Display','Date & Time','Product Name','End User Price']].rename(columns={'Phone_Display':'Phone','End User Price':'Client price (NIS)'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.info("No refunds")
            with tabs[7]:
                if len(result['failed']) > 0:
                    show = result['failed'][['Operator','Phone_Display','Date & Time','Product Name','Error description']].rename(columns={'Phone_Display':'Phone'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.success("✅ No failed!")

            st.markdown("---")
            st.markdown("### 💰 Net Billing Summary")
            net_data = {
                'Item': ['Client charges — DONE+CANCELLED','Refunds (credit back)','PENDING',
                         'NET Our Total','Supplier CBD (matched)','Supplier Only CBD','📊 Real Gap'],
                'Partner (NIS)': [round(t['partner_eup'],2), round(t['partner_ref'],2), round(t['pending_eup'],2),
                                  round(t['partner_eup']+t['partner_ref'],2), '—','—','—'],
                '012Talk (NIS)': [round(t['talk012_eup'],2), round(t['talk012_ref'],2), 0,
                                  round(t['talk012_eup']+t['talk012_ref'],2), '—','—','—'],
                'TOTAL (NIS)': [
                    round(t['partner_eup']+t['talk012_eup'],2),
                    round(t['partner_ref']+t['talk012_ref'],2),
                    round(t['pending_eup'],2),
                    round(t['partner_eup']+t['talk012_eup']+t['partner_ref']+t['talk012_ref'],2),
                    round(t['sup_cbd'],2), round(t['sup_only_cbd'],2), round(t['real_gap'],2)
                ],
            }
            st.dataframe(pd.DataFrame(net_data), use_container_width=True, hide_index=True)

            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 Save to Monthly History", use_container_width=True, key="pt_save"):
                    if check_sheets_banner():
                        st.session_state['pt_do_save'] = True
            with col2:
                excel_buf = create_excel_report(result, rdate, 'Partner & 012Talk')
                st.download_button("📥 Download Excel Report", data=excel_buf,
                    file_name=f"Partner_012Talk_{rdate.replace('-','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary")

            # Confirmation dialog and actual save — outside columns so it renders full-width
            if st.session_state.get('pt_do_save'):
                if save_confirmation_ui(rdate, 'partner', 'pt'):
                    st.session_state.pop('pt_do_save', None)
                    record = {
                        'date': rdate, 'operator_tab': 'partner',
                        'sup_cbd': round(t['sup_cbd'],2), 'our_eup': round(t['our_eup'],2),
                        'diff': round(t['diff'],2), 'matched_count': t['matched_count'],
                        'sup_only_count': t['sup_only_count'], 'sup_only_cbd': round(t['sup_only_cbd'],2),
                        'our_only_count': t['our_only_count'], 'our_only_eup': round(t['our_only_eup'],2),
                        'real_gap': round(t['real_gap'],2), 'pending_count': t['pending_count'],
                        'refunds_eup': round(t['refunds_eup'],2),
                        'net_billed': round(t['partner_eup']+t['talk012_eup']+t['partner_ref']+t['talk012_ref'],2),
                    }
                    ok, msg = save_to_sheets(record)
                    detail_rows = build_detail_rows(rdate, 'partner', result['sup_only'], result['our_only'])
                    ok2, msg2 = save_details_to_sheets(rdate, 'partner', detail_rows)
                    if ok:
                        st.cache_data.clear()
                        st.success(f"✅ {msg}")
                    else:
                        st.warning(f"⚠️ {msg}")
                    if ok2:
                        st.info(f"📋 {msg2}")
                    else:
                        st.warning(f"⚠️ Details: {msg2}")
                    if ok:
                        _send_daily_email('Partner & 012Talk', rdate, result, t)

    # ============================================================
    # PAGE: PELEPHONE
    # ============================================================
    elif page == "⭐ Pelephone Reconciliation":
        render_header(
            "Pelephone Reconciliation",
            "Supplier vs Our System (Pelephone + GlobalSim + eSIM)",
            [LOGO_PAYX, LOGO_PELE]
        )
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown("**1️⃣ Supplier File (.xlsx)**")
            sup_file = st.file_uploader("Pelephone Supplier", type=['xlsx','xls'],
                                        label_visibility="collapsed", key="pe_sup")
            st.caption("Pelephone supplier report (col F = price)")
        with col2:
            st.markdown("**2️⃣ Pelephone EPRS File (.csv)**")
            pele_file = st.file_uploader("Pelephone EPRS", type=['csv'],
                                         label_visibility="collapsed", key="pe_pele")
        with col3:
            st.markdown("**3️⃣ GlobalSim EPRS File (.csv)**")
            glob_file = st.file_uploader("GlobalSim EPRS", type=['csv'],
                                         label_visibility="collapsed", key="pe_glob")
        with col4:
            st.markdown("**4️⃣ eSIM EPRS File (.csv)**")
            esim_file = st.file_uploader("eSIM EPRS", type=['csv'],
                                         label_visibility="collapsed", key="pe_esim")
            st.caption("eSIM: our 5 NIS vs supplier 7.67 NIS")

        if sup_file and pele_file:
            if st.button("▶ Run Reconciliation", type="primary", use_container_width=True, key="pe_run"):
                with st.spinner("Processing..."):
                    sup_df,  e1 = load_supplier_pelephone(sup_file.read())
                    pele_df, e2 = load_our(pele_file.read(), 'Pelephone')
                    if e1: st.error(f"Supplier error: {e1}"); return
                    if e2: st.error(f"Pelephone error: {e2}"); return
                    glob_df, e3 = load_our(glob_file.read(), 'GlobalSim') if glob_file else (pd.DataFrame(), None)
                    esim_df, e4 = load_our(esim_file.read(), 'eSIM') if esim_file else (pd.DataFrame(), None)
                    if e3: st.error(f"GlobalSim error: {e3}"); return
                    if e4: st.error(f"eSIM error: {e4}"); return

                    auto_date = date.today().strftime('%Y-%m-%d')
                    if 'Sup_Date' in sup_df.columns and len(sup_df) > 0:
                        try:
                            raw = str(sup_df['Sup_Date'].iloc[0])
                            auto_date = pd.to_datetime(raw, dayfirst=True).strftime('%Y-%m-%d')
                            st.info(f"📅 Date detected: **{auto_date}**")
                        except: pass

                    result = run_recon_pelephone(sup_df, pele_df, glob_df, esim_df, auto_date)
                    shifts_our, shifts_sup = cross_day_match(result, auto_date, 'pelephone')
                    st.session_state['pe_result'] = result
                    st.session_state['pe_date'] = auto_date
                    st.session_state['pe_shifts_our'] = shifts_our
                    st.session_state['pe_shifts_sup'] = shifts_sup
                    st.success("✅ Complete!")

        if 'pe_result' in st.session_state:
            result = st.session_state['pe_result']
            t = result['totals']
            rdate = st.session_state['pe_date']
            shifts_our = st.session_state.get('pe_shifts_our',[])
            shifts_sup = st.session_state.get('pe_shifts_sup',[])

            st.markdown("---")
            c1,c2,c3,c4,c5 = st.columns(5)
            c1.metric("✅ Matched", f"{t['matched_count']:,}")
            c2.metric("❌ Supplier Only", t['sup_only_count'],
                      delta=f"{t['sup_only_price']:,.2f} NIS", delta_color="inverse")
            c3.metric("⚠️ Our Only", t['our_only_count'],
                      delta=f"{t['our_only_eup']:,.2f} NIS", delta_color="inverse")
            c4.metric("📟 eSIM", t.get('esim_count', 0),
                      delta=f"our {t.get('esim_our_total', 0):,.2f} / sup {t.get('esim_sup_total', 0):,.2f} NIS",
                      delta_color="off")
            gap = t['real_gap']
            c5.metric("📊 Real Gap",
                      f"+{gap:,.2f} NIS" if gap>0 else (f"{gap:,.2f} NIS" if gap<0 else "0.00 NIS ✅"),
                      delta_color="inverse" if gap>0 else "normal")

            if shifts_our:
                st.success(f"✅ Date Shift: {len(shifts_our)} phone(s) confirmed — {', '.join(shifts_our)}")

            render_action_required(result['sup_only'], result['our_only'], rdate, shifts_our, shifts_sup)

            st.markdown("---")
            tabs = st.tabs([
                "📋 Summary",
                f"❌ Supplier Only ({t['sup_only_count']})",
                f"⚠️ Our Only ({t['our_only_count']})",
                f"✅ Matched ({t['matched_count']})",
                f"↩️ Refunds ({t['refunds_count']})",
                f"🔴 Failed ({t['failed_count']})",
            ])
            with tabs[0]:
                render_summary_tab(result['sup_only'], result['our_only'], t, rdate, 'pelephone')
                if t.get('esim_count', 0) > 0:
                    st.markdown("#### 📟 eSIM Summary")
                    _es_diff = round(t.get('esim_sup_total', 0) - t.get('esim_our_total', 0), 2)
                    st.dataframe(pd.DataFrame({
                        'Item': ['eSIM transactions',
                                 'Our client price total (NIS)',
                                 'Supplier price total (NIS)',
                                 'Price difference — supplier minus ours (NIS)'],
                        'Value': [t.get('esim_count', 0),
                                  f"{t.get('esim_our_total', 0):,.2f}",
                                  f"{t.get('esim_sup_total', 0):,.2f}",
                                  f"{_es_diff:,.2f}"],
                    }), use_container_width=True, hide_index=True)
                    st.caption("Expected difference is +2.67 NIS per eSIM (supplier 7.67 vs our 5.00).")
                if t['unexpected_diff_count'] > 0:
                    st.warning(f"⚠️ {t['unexpected_diff_count']} transaction(s) with unexpected price differences — check Matched tab")
            with tabs[1]:
                if len(result['sup_only']) > 0:
                    sup_only_cols = ['Phone_Display','Sup_Date','TOPUP_ITEM','TOPUP_PRICE','Check_Instruction']
                    if 'Sup_Time' in result['sup_only'].columns:
                        sup_only_cols.insert(2, 'Sup_Time')
                    show = result['sup_only'][[c for c in sup_only_cols if c in result['sup_only'].columns]].copy()
                    show.columns = show.columns.str.replace('Phone_Display','Phone').str.replace('Sup_Date','Date').str.replace('Sup_Time','Time').str.replace('TOPUP_ITEM','Product').str.replace('TOPUP_PRICE','Price (NIS)').str.replace('Check_Instruction','What to check')
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.success("✅ No supplier-only!")
            with tabs[2]:
                if len(result['our_only']) > 0:
                    show = result['our_only'][['Phone_Display','Date & Time','Operator','Product Name','End User Price','Check_Instruction']].rename(
                        columns={'Phone_Display':'Phone','End User Price':'Client price (NIS)','Check_Instruction':'What to check'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.success("✅ No our-only!")
            with tabs[3]:
                if len(result['matched']) > 0:
                    st.dataframe(result['matched'].rename(
                        columns={'Our EUP (NIS)': 'Our Client Price (NIS)'}),
                        use_container_width=True, hide_index=True)
                else: st.info("No matched records")
            with tabs[4]:
                if len(result['refunds']) > 0:
                    show = result['refunds'][['Operator','Phone_Display','Date & Time','Product Name','End User Price']].rename(
                        columns={'Phone_Display':'Phone','End User Price':'Client price (NIS)'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.info("No refunds")
            with tabs[5]:
                if len(result['failed']) > 0:
                    show = result['failed'][['Operator','Phone_Display','Date & Time','Product Name','Error description']].rename(
                        columns={'Phone_Display':'Phone'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.success("✅ No failed!")

            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 Save to Monthly History", use_container_width=True, key="pe_save"):
                    if check_sheets_banner():
                        st.session_state['pe_do_save'] = True
            with col2:
                excel_buf = create_excel_report(result, rdate, 'Pelephone')
                st.download_button("📥 Download Excel Report", data=excel_buf,
                    file_name=f"Pelephone_{rdate.replace('-','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary")

            if st.session_state.get('pe_do_save'):
                if save_confirmation_ui(rdate, 'pelephone', 'pe'):
                    st.session_state.pop('pe_do_save', None)
                    record = {
                        'date': rdate, 'operator_tab': 'pelephone',
                        'sup_cbd': round(t['sup_price'],2), 'our_eup': round(t['our_eup'],2),
                        'diff': round(t['diff'],2), 'matched_count': t['matched_count'],
                        'sup_only_count': t['sup_only_count'], 'sup_only_cbd': round(t['sup_only_price'],2),
                        'our_only_count': t['our_only_count'], 'our_only_eup': round(t['our_only_eup'],2),
                        'real_gap': round(t['real_gap'],2), 'pending_count': 0,
                        'refunds_eup': round(t['refunds_eup'],2), 'net_billed': round(t['our_eup'],2),
                        'esim_count': t.get('esim_count', 0),
                        'esim_our_total': round(t.get('esim_our_total', 0), 2),
                        'esim_sup_total': round(t.get('esim_sup_total', 0), 2),
                    }
                    ok, msg = save_to_sheets(record)
                    detail_rows = build_detail_rows(rdate, 'pelephone', result['sup_only'], result['our_only'])
                    ok2, msg2 = save_details_to_sheets(rdate, 'pelephone', detail_rows)
                    if ok:
                        st.success(f"✅ {msg}")
                    else:
                        st.warning(f"⚠️ {msg}")
                    if ok2:
                        st.info(f"📋 {msg2}")
                    else:
                        st.warning(f"⚠️ Details: {msg2}")
                    if ok:
                        _send_daily_email('Pelephone', rdate, result, t)

    # ============================================================
    # PAGE: CELLCOM
    # ============================================================
    elif page == "📡 Cellcom Reconciliation":
        render_header(
            "Cellcom Reconciliation",
            "Supplier vs Our System (Cellcom)",
            [LOGO_PAYX, LOGO_CELL]
        )
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**1️⃣ Cellcom Supplier File (.xlsx)**")
            sup_file = st.file_uploader("Cellcom Supplier", type=['xlsx','xls'],
                                        label_visibility="collapsed", key="ce_sup")
            st.caption("Supplier report — col C = charge amount, col G = phone")
        with col2:
            st.markdown("**2️⃣ Cellcom EPRS File (.csv)**")
            our_file = st.file_uploader("Cellcom EPRS", type=['csv'],
                                        label_visibility="collapsed", key="ce_our")
            st.caption("Our system export — Cellcom operator")

        if sup_file and our_file:
            if st.button("▶ Run Reconciliation", type="primary", use_container_width=True, key="ce_run"):
                with st.spinner("Processing..."):
                    sup_df, e1 = load_supplier_cellcom(sup_file.read())
                    our_df, e2 = load_our(our_file.read(), 'Cellcom')
                    if e1: st.error(f"Supplier error: {e1}"); return
                    if e2: st.error(f"Our file error: {e2}"); return

                    auto_date = date.today().strftime('%Y-%m-%d')
                    if 'Sup_Date' in sup_df.columns and len(sup_df) > 0:
                        try:
                            auto_date = sup_df['Sup_Date'].iloc[0].strftime('%Y-%m-%d')
                            st.info(f"📅 Date detected: **{auto_date}**")
                        except: pass

                    result = run_recon_cellcom(sup_df, our_df, auto_date)
                    shifts_our, shifts_sup = cross_day_match(result, auto_date, 'cellcom')
                    st.session_state['ce_result'] = result
                    st.session_state['ce_date'] = auto_date
                    st.session_state['ce_shifts_our'] = shifts_our
                    st.session_state['ce_shifts_sup'] = shifts_sup
                    st.success("✅ Complete!")

        if 'ce_result' in st.session_state:
            result = st.session_state['ce_result']
            t = result['totals']
            rdate = st.session_state['ce_date']
            shifts_our = st.session_state.get('ce_shifts_our',[])
            shifts_sup = st.session_state.get('ce_shifts_sup',[])

            st.markdown("---")
            c1,c2,c3,c4,c5 = st.columns(5)
            c1.metric("✅ Matched", f"{t['matched_count']:,}")
            c2.metric("❌ Supplier Only", t['sup_only_count'],
                      delta=f"{t['sup_only_cbd']:,.2f} NIS", delta_color="inverse")
            c3.metric("⚠️ Our Only", t['our_only_count'],
                      delta=f"{t['our_only_eup']:,.2f} NIS", delta_color="inverse")
            c4.metric("💱 Expected Discount", f"{t['total_expected_discount']:,.2f} NIS",
                      delta=f"~5 NIS × {t['matched_count']} (fixed tariffs excluded)", delta_color="off")
            c5.metric("⚠️ Unexplained Diff",
                      f"{t['unexplained_diff']:,.2f} NIS" if abs(t['unexplained_diff']) > 0.01 else "0.00 ✅",
                      delta="Should be 0" if abs(t['unexplained_diff']) > 0.01 else None,
                      delta_color="inverse" if abs(t['unexplained_diff']) > 0.01 else "off")

            if shifts_our:
                st.success(f"✅ Date Shift: {len(shifts_our)} phone(s) confirmed — {', '.join(shifts_our)}")

            if abs(t['unexplained_diff']) > 0.01:
                st.warning(f"⚠️ Unexplained price difference: {t['unexplained_diff']:,.2f} NIS — "
                          f"Expected discount: {t['total_expected_discount']:,.2f} NIS but actual differs. "
                          f"Check '{t['price_anomaly_count']}' anomaly transactions in Matched tab.")

            render_action_required(result['sup_only'], result['our_only'], rdate, shifts_our, shifts_sup)

            st.markdown("---")
            tabs = st.tabs([
                "📋 Summary",
                f"❌ Supplier Only ({t['sup_only_count']})",
                f"⚠️ Our Only ({t['our_only_count']})",
                f"✅ Matched ({t['matched_count']})",
                f"↩️ Refunds ({t['refunds_count']})",
                f"🔴 Failed ({t['failed_count']})",
            ])
            with tabs[0]:
                render_summary_tab(result['sup_only'], result['our_only'], t, rdate, 'cellcom')
                st.markdown("#### 💱 Price Analysis")
                price_data = {
                    'Item': [
                        f'Matched transactions ({t["matched_count"]})',
                        'Expected total discount (~5 NIS each, excl. fixed tariffs)',
                        'Actual total discount',
                        '⚠️ Unexplained difference',
                    ],
                    'Amount (NIS)': [
                        round(t['our_eup'],2),
                        round(t['total_expected_discount'],2),
                        round(t['our_eup'] - t['sup_cbd'],2),
                        round(t['unexplained_diff'],2),
                    ]
                }
                st.dataframe(pd.DataFrame(price_data), use_container_width=True, hide_index=True)
            with tabs[1]:
                if len(result['sup_only']) > 0:
                    show = result['sup_only'][['Phone_Display','Sup_Date','CBD','Check_Instruction']].rename(
                        columns={'Phone_Display':'Phone','Sup_Date':'Date','CBD':'CBD (NIS)','Check_Instruction':'What to check'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.success("✅ No supplier-only!")
            with tabs[2]:
                if len(result['our_only']) > 0:
                    show = result['our_only'][['Phone_Display','Date & Time','Product Name','End User Price','Check_Instruction']].rename(
                        columns={'Phone_Display':'Phone','End User Price':'Client price (NIS)','Check_Instruction':'What to check'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.success("✅ No our-only!")
            with tabs[3]:
                if len(result['matched']) > 0:
                    st.dataframe(result['matched'].rename(
                        columns={'Our EUP (NIS)': 'Our Client Price (NIS)'}),
                        use_container_width=True, hide_index=True)
                    if len(result['anomalies']) > 0:
                        st.warning(f"⚠️ {len(result['anomalies'])} transaction(s) with unexpected price diff:")
                        st.dataframe(result['anomalies'], use_container_width=True, hide_index=True)
                else: st.info("No matched records")
            with tabs[4]:
                if len(result['refunds']) > 0:
                    show = result['refunds'][['Phone_Display','Date & Time','Product Name','End User Price']].rename(
                        columns={'Phone_Display':'Phone','End User Price':'Client price (NIS)'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.info("No refunds")
            with tabs[5]:
                if len(result['failed']) > 0:
                    show = result['failed'][['Phone_Display','Date & Time','Product Name','Error description']].rename(
                        columns={'Phone_Display':'Phone'})
                    st.dataframe(show, use_container_width=True, hide_index=True)
                else: st.success("✅ No failed!")

            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 Save to Monthly History", use_container_width=True, key="ce_save"):
                    if check_sheets_banner():
                        st.session_state['ce_do_save'] = True
            with col2:
                excel_buf = create_excel_report(result, rdate, 'Cellcom')
                st.download_button("📥 Download Excel Report", data=excel_buf,
                    file_name=f"Cellcom_{rdate.replace('-','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary")

            if st.session_state.get('ce_do_save'):
                if save_confirmation_ui(rdate, 'cellcom', 'ce'):
                    st.session_state.pop('ce_do_save', None)
                    record = {
                        'date': rdate, 'operator_tab': 'cellcom',
                        'sup_cbd': round(t['sup_cbd'],2), 'our_eup': round(t['our_eup'],2),
                        'diff': round(t['our_eup']-t['sup_cbd'],2),
                        'matched_count': t['matched_count'],
                        'sup_only_count': t['sup_only_count'], 'sup_only_cbd': round(t['sup_only_cbd'],2),
                        'our_only_count': t['our_only_count'], 'our_only_eup': round(t['our_only_eup'],2),
                        'real_gap': round(t['real_gap'],2), 'pending_count': t['pending_count'],
                        'refunds_eup': round(t['refunds_eup'],2), 'net_billed': round(t['our_eup'],2),
                    }
                    ok, msg = save_to_sheets(record)
                    detail_rows = build_detail_rows(rdate, 'cellcom', result['sup_only'], result['our_only'])
                    ok2, msg2 = save_details_to_sheets(rdate, 'cellcom', detail_rows)
                    if ok:
                        st.success(f"✅ {msg}")
                    else:
                        st.warning(f"⚠️ {msg}")
                    if ok2:
                        st.info(f"📋 {msg2}")
                    else:
                        st.warning(f"⚠️ Details: {msg2}")
                    if ok:
                        _send_daily_email('Cellcom', rdate, result, t)

    # ============================================================
    # PAGE: MONTHLY SUMMARY
    # ============================================================
    elif page == "🚌 RavKav Reconciliation":
        render_header("RavKav Reconciliation",
                      "Daily totals check + uid-level investigation",
                      [LOGO_PAYX], extra_labels=['RavKav'])
        st.caption("Client price (NIS) = the amount paid by the end client "
                   "(column 'End User Price' in our system export).")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("**1️⃣ Our System Logs (.csv)**")
            rk_our_file = st.file_uploader("RavKav ours", type=['csv'],
                                           label_visibility="collapsed", key="rk_our")
            st.caption("Export from our system — RavKav operator")
        with col2:
            st.markdown("**2️⃣ RavKav Daily Totals (.csv)**")
            rk_tot_file = st.file_uploader("RavKav totals", type=['csv'],
                                           label_visibility="collapsed", key="rk_tot")
            st.caption("One-line totals report (amounts in agorot)")
        with col3:
            st.markdown("**3️⃣ RavKav Detailed Report (.xlsx) — optional**")
            rk_det_file = st.file_uploader("RavKav detailed", type=['xlsx', 'xls'],
                                           label_visibility="collapsed", key="rk_det")
            st.caption("Request from RavKav when totals do not match")

        if rk_our_file and (rk_tot_file or rk_det_file):
            if st.button("▶ Run Reconciliation", type="primary",
                         use_container_width=True, key="rk_run"):
                with st.spinner("Processing..."):
                    our_df, e1 = load_ravkav_our(rk_our_file.read())
                    if e1: st.error(f"Our file error: {e1}"); return
                    sup_totals, sup_det = None, None
                    if rk_tot_file:
                        sup_totals, e2 = load_ravkav_totals(rk_tot_file.read())
                        if e2: st.error(f"Totals file error: {e2}"); return
                    if rk_det_file:
                        sup_det, e3 = load_ravkav_detailed(rk_det_file.read())
                        if e3: st.error(f"Detailed file error: {e3}"); return
                    auto_date = date.today().strftime('%Y-%m-%d')
                    try:
                        _dts = our_df['_dt'].dropna()
                        if len(_dts) > 0:
                            auto_date = _dts.max().strftime('%Y-%m-%d')
                            st.info(f"📅 Date detected: **{auto_date}**")
                    except Exception:
                        pass
                    result = run_recon_ravkav(our_df, sup_totals, sup_det, auto_date)
                    st.session_state['rk_result'] = result
                    st.session_state['rk_date'] = auto_date
                    st.success("✅ Complete!")

        if 'rk_result' in st.session_state:
            result = st.session_state['rk_result']
            t = result['totals']
            rdate = st.session_state['rk_date']

            st.markdown("---")
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("✅ Charged (ours)", t['done_count'],
                      delta=f"{t['our_charged']:,.2f} NIS", delta_color="off")
            c2.metric("🔴 Failed (ours)", t['failed_count'])
            c3.metric("↩️ RavKav-refunded", t['cancelled_count'],
                      delta=f"{t['our_refunded']:,.2f} NIS", delta_color="off")
            if t.get('sup_charged') is not None:
                c4.metric("RavKav charged", f"{t['sup_charged']:,.2f} NIS",
                          delta=f"refunded {t['sup_refunded']:,.2f} NIS", delta_color="off")
                c5.metric("📊 Gap", f"{t['gap_net']:,.2f} NIS",
                          delta="RavKav net minus our net",
                          delta_color="inverse" if abs(t['gap_net']) > 0.009 else "off")

            if t.get('client_ref_count'):
                st.caption(f"🧾 Our client refunds (internal, no RavKav uid): "
                           f"{t['client_ref_count']} row(s), {t['client_ref_sum']:,.2f} NIS — "
                           f"included in net billing, excluded from the RavKav comparison.")

            if t.get('sup_charged') is not None:
                if abs(t['gap_charged']) < 0.01 and abs(t['gap_refunded']) < 0.01:
                    st.success("✅ Amounts match — charges and refunds are identical on both sides.")
                else:
                    if abs(t['gap_charged']) >= 0.01:
                        st.error(f"⚠️ Charged amounts differ by {t['gap_charged']:,.2f} NIS "
                                 f"(ours {t['our_charged']:,.2f} vs RavKav {t['sup_charged']:,.2f})")
                    if abs(t['gap_refunded']) >= 0.01:
                        st.error(f"⚠️ Refunded amounts differ by {t['gap_refunded']:,.2f} NIS "
                                 f"(ours {t['our_refunded']:,.2f} vs RavKav {t['sup_refunded']:,.2f})")
                    if t['adj_count']:
                        st.warning(f"🧾 Manual adjustments in our log (no RavKav uid): "
                                   f"{t['adj_count']} row(s), {t['adj_sum']:,.2f} NIS — "
                                   f"part of the difference")
                    st.info("Request the detailed report from RavKav and upload it "
                            "as file 3️⃣ for a uid-level breakdown.")

            if result.get('detailed') is not None:
                d = result['detailed']
                st.markdown("### 🔬 Detailed uid-level comparison")
                dd1, dd2, dd3, dd4 = st.columns(4)
                dd1.metric("✅ Matching uids", d['matched_count'])
                dd2.metric("⚠️ Failed here / charged there", len(d['failed_but_charged']))
                dd3.metric("Only in our log", len(d['only_ours']))
                dd4.metric("Only at RavKav", len(d['only_theirs']))
                for _title, _key in [
                        ("⚠️ FAILED in our system but CHARGED by RavKav — main suspects", 'failed_but_charged'),
                        ("DONE in our system but failed at RavKav", 'done_not_charged'),
                        ("Amount mismatches", 'amount_mismatch'),
                        ("Only in our log (uid unknown to RavKav)", 'only_ours'),
                        ("Only at RavKav", 'only_theirs'),
                        ("🧾 Manual adjustments (no uid)", 'adjustments')]:
                    _dfx = d[_key]
                    if len(_dfx) > 0:
                        st.markdown(f"**{_title} ({len(_dfx)})**")
                        st.dataframe(_dfx, use_container_width=True, hide_index=True)

                render_action_required(result['sup_only'], result['our_only'], rdate)

            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 Save to Monthly History", use_container_width=True, key="rk_save"):
                    if check_sheets_banner():
                        st.session_state['rk_do_save'] = True
            with col2:
                excel_buf = create_excel_report(result, rdate, 'RavKav')
                st.download_button("📥 Download Excel Report", data=excel_buf,
                    file_name=f"RavKav_{rdate.replace('-','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary")

            if st.session_state.get('rk_do_save'):
                if save_confirmation_ui(rdate, 'ravkav', 'rk'):
                    st.session_state.pop('rk_do_save', None)
                    _sup_cbd = t['sup_charged'] if t.get('sup_charged') is not None \
                        else round(t['our_charged'] + t['real_gap'], 2)
                    record = {
                        'date': rdate, 'operator_tab': 'ravkav',
                        'sup_cbd': round(_sup_cbd, 2), 'our_eup': round(t['our_charged'], 2),
                        'diff': round(t['gap_charged'], 2),
                        'matched_count': t['done_count'],
                        'sup_only_count': t['sup_only_count'],
                        'sup_only_cbd': round(t['sup_only_cbd'], 2),
                        'our_only_count': t['our_only_count'],
                        'our_only_eup': round(t['our_only_eup'], 2),
                        'real_gap': round(t['real_gap'], 2), 'pending_count': 0,
                        'refunds_eup': round(t['refunds_eup'], 2),
                        'net_billed': round(t['our_charged'] + t.get('client_ref_sum', 0), 2),
                    }
                    ok, msg = save_to_sheets(record)
                    detail_rows = build_detail_rows(rdate, 'ravkav',
                                                    result['sup_only'], result['our_only'])
                    ok2, msg2 = save_details_to_sheets(rdate, 'ravkav', detail_rows)
                    if ok:
                        st.cache_data.clear()
                        st.success(f"✅ {msg}")
                    else:
                        st.warning(f"⚠️ {msg}")
                    if ok2:
                        st.info(f"📋 {msg2}")
                    else:
                        st.warning(f"⚠️ Details: {msg2}")
                    if ok:
                        _send_daily_email('RavKav', rdate, result, t)

    elif page == "📅 Monthly Summary":
        render_header("Monthly Summary", "All operators — month overview", [LOGO_PAYX])
        op_filter_pre = st.selectbox("Operator", ["partner","pelephone","cellcom","ravkav"], key="op_pre")
        sh = get_spreadsheet(op_filter_pre)
        available_months = []
        if sh is not None:
            try:
                for ws in sh.worksheets():
                    if _is_month_sheet(ws.title):
                        try:
                            dt = datetime.strptime(ws.title, '%B %Y')
                            available_months.append(dt.strftime('%Y-%m'))
                        except: pass
                available_months = sorted(set(available_months), reverse=True)
            except: pass

        if not available_months:
            history = _load_local_history()
            available_months = sorted(set(h['date'][:7] for h in history), reverse=True)

        if not available_months:
            st.info("No history yet. Run daily reconciliations and click 'Save to Monthly History'.")
            return

        selected_month = st.selectbox("Select Month", available_months,
            format_func=lambda m: datetime.strptime(m, '%Y-%m').strftime('%B %Y'))

        month_history = load_history(month=selected_month, operator_tab=op_filter_pre)
        if not month_history:
            st.warning("No data for selected month/operator")
            return

        # De-duplicate: if same date has multiple rows (from safe appends), show only latest
        seen = {}
        for rec in month_history:
            key = rec.get('date','')
            seen[key] = rec  # last one wins — that's the most recent run
        month_history_deduped = list(seen.values())

        total_sup  = sum(h.get('sup_cbd',0) for h in month_history_deduped)
        total_eup  = sum(h.get('our_eup',0) for h in month_history_deduped)
        total_gap  = sum(h.get('real_gap',0) for h in month_history_deduped)
        total_ref  = sum(h.get('refunds_eup',0) for h in month_history_deduped)

        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("📅 Days", len(month_history_deduped))
        c2.metric("Supplier Total", f"{total_sup:,.2f} NIS")
        c3.metric("Client Charges Total", f"{total_eup:,.2f} NIS")
        c4.metric("↩️ Refunds", f"{total_ref:,.2f} NIS")
        c5.metric("📊 Monthly Real Gap", f"{total_gap:,.2f} NIS",
                  delta_color="inverse" if total_gap>0 else "normal")

        st.markdown("---")
        st.markdown("### 🔍 Verification status (live)")
        _det = load_month_details(op_filter_pre, selected_month)

        def _famt(v):
            try:
                return float(str(v).replace(',', '') or 0)
            except Exception:
                return 0.0

        _res_ok = [r for r in _det if str(r.get('verified', '')).startswith('✅')]
        _dups   = [r for r in _det if str(r.get('verified', '')).startswith('🔵')]
        _open   = [r for r in _det if str(r.get('verified', '')).startswith('⬜')
                   or str(r.get('verified', '')).startswith('❌')]
        _rem_sup = sum(_famt(r.get('amount')) for r in _open
                       if str(r.get('category', '')) == 'Supplier Only')
        _rem_our = sum(_famt(r.get('amount')) for r in _open
                       if str(r.get('category', '')) == 'Our Only')
        _remaining_gap = round(_rem_sup - _rem_our, 2)

        d1, d2, d3, d4 = st.columns(4)
        d1.metric("✅ Resolved", len(_res_ok),
                  delta=f"{sum(_famt(r.get('amount')) for r in _res_ok):,.2f} NIS",
                  delta_color="off")
        d2.metric("🔵 Duplicates → Refunds", len(_dups),
                  delta=f"{sum(_famt(r.get('amount')) for r in _dups):,.2f} NIS",
                  delta_color="off")
        d3.metric("Open (⬜ + ❌)", len(_open),
                  delta=f"Sup {_rem_sup:,.2f} / Our {_rem_our:,.2f} NIS",
                  delta_color="off")
        d4.metric("📊 Remaining Gap", f"{_remaining_gap:,.2f} NIS",
                  delta="only ⬜ and ❌ remain here",
                  delta_color="inverse" if _remaining_gap > 0 else "normal")
        if _open:
            with st.expander(f"🔎 Open items ({len(_open)}) — where the remaining gap comes from"):
                _odf = pd.DataFrame(_open)
                _ocols = [c for c in ['date', 'category', 'phone', 'product', 'amount',
                                      'verified', 'verified_by', 'verified_at']
                          if c in _odf.columns]
                st.dataframe(_odf[_ocols], use_container_width=True, hide_index=True)
        st.caption("Monthly Real Gap above is frozen at save time. "
                   "Remaining Gap reflects current verification statuses "
                   "(updates within ~2 minutes after saving a verification).")

        st.markdown("---")
        df = pd.DataFrame(month_history_deduped)
        if op_filter_pre == 'pelephone':
            def _ecol(_n):
                if _n in df.columns:
                    return pd.to_numeric(df[_n], errors='coerce').fillna(0)
                return pd.Series([0.0] * len(df), index=df.index)
            df['esim_count'] = _ecol('esim_count').astype(int)
            df['esim_price_diff'] = (_ecol('esim_sup_total') - _ecol('esim_our_total')).round(2)
        show_cols = [c for c in ['date','operator_tab','matched_count','sup_cbd','our_eup',
                                  'diff','real_gap','esim_count','esim_price_diff',
                                  'refunds_eup','net_billed'] if c in df.columns]
        _disp = df[show_cols].copy()
        _tot = {c: '' for c in show_cols}
        _tot['date'] = '📊 TOTAL'
        if 'operator_tab' in _tot:
            _tot['operator_tab'] = op_filter_pre
        for _c in ['matched_count', 'sup_cbd', 'our_eup', 'diff', 'real_gap',
                   'esim_count', 'esim_price_diff', 'refunds_eup', 'net_billed']:
            if _c in show_cols:
                _tot[_c] = round(float(pd.to_numeric(df[_c], errors='coerce').fillna(0).sum()), 2)
        _disp = pd.concat([_disp, pd.DataFrame([_tot])], ignore_index=True)
        st.dataframe(_disp, use_container_width=True, hide_index=True)
        if op_filter_pre == 'pelephone' and 'esim_count' in df.columns:
            _en = pd.to_numeric(df['esim_count'], errors='coerce').fillna(0).sum()
            _eo = pd.to_numeric(df.get('esim_our_total'), errors='coerce').fillna(0).sum()
            _esu = pd.to_numeric(df.get('esim_sup_total'), errors='coerce').fillna(0).sum()
            st.caption(f"📟 eSIM this month: {int(_en)} transactions | "
                       f"our {_eo:,.2f} NIS | supplier {_esu:,.2f} NIS | "
                       f"price diff {_esu - _eo:,.2f} NIS")

        month_label = datetime.strptime(selected_month, '%Y-%m').strftime('%B %Y')
        st.download_button(
            f"📥 Download Monthly Report — {month_label}",
            data=create_monthly_excel(month_history_deduped, month_label),
            file_name=f"Monthly_{selected_month.replace('-','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary"
        )

    # ============================================================
    # PAGE: PENDING VERIFICATION
    # ============================================================
    elif page == "⏳ Pending Verification":
        render_header("Pending Verification", "Phones awaiting manual check", [LOGO_PAYX])

        col_f, col_r = st.columns([4, 1])
        with col_r:
            if st.button("🔄 Refresh", use_container_width=True, key="refresh_btn"):
                st.session_state.pop('pending_local', None)
                st.rerun()
        with col_f:
            saved_filter = st.session_state.get('pending_op_filter_saved', 'All')
            _op_opts = ["All", "partner", "pelephone", "cellcom", "ravkav"]
            op_filter = st.selectbox(
                "Filter by operator:",
                _op_opts,
                index=_op_opts.index(saved_filter) if saved_filter in _op_opts else 0,
                key="pending_op_filter"
            )
            st.session_state['pending_op_filter_saved'] = op_filter

        phone_search = st.text_input("🔍 Search by phone number:", key="pend_phone_search", placeholder="e.g. 0541234567")

        # Show result of the previous save (survives st.rerun)
        if 'pend_last_result' in st.session_state:
            for _m in st.session_state.pop('pend_last_result'):
                if _m.startswith('✅'):
                    st.success(_m)
                else:
                    st.error(_m)

        # Surface per-operator load problems instead of hiding them
        for _e in st.session_state.get('pending_load_errors', []):
            st.warning(f"⚠️ Load issue — {_e}")

        # Drop cache written by an older app version (rows have no _row)
        _pl = st.session_state.get('pending_local')
        if _pl and '_row' not in _pl[0]:
            st.session_state.pop('pending_local', None)

        # Load once per page visit, search/filter locally
        if 'pending_local' not in st.session_state:
            with st.spinner("Loading pending verifications..."):
                loaded = load_pending_verifications()
                if loaded:
                    st.session_state['pending_local'] = loaded
                    st.session_state['pending_load_ok'] = True
                else:
                    sh_test = get_spreadsheet()
                    if sh_test is None:
                        st.session_state['pending_load_ok'] = False
                    else:
                        st.session_state['pending_load_ok'] = True
                    st.session_state['pending_local'] = loaded

        all_pending = st.session_state.get('pending_local', [])
        load_ok = st.session_state.get('pending_load_ok', True)

        if not load_ok and not all_pending:
            st.error("❌ Google Sheets unavailable. Click 🔄 Refresh to retry.")
            return

        if not all_pending:
            st.success("✅ Nothing pending — all verified!")
            return

        pending = all_pending
        if op_filter != "All":
            pending = [r for r in pending if r.get('operator_tab', '') == op_filter]
        if phone_search:
            search_norm = phone_search.strip().lstrip('0')
            pending = [r for r in pending if search_norm in str(r.get('phone', '')).replace('.0', '').lstrip('0')]

        st.warning(f"⏳ {len(pending)} phone(s) shown — {len(all_pending)} total pending across all operators")

        if not pending:
            st.info("No results for current filter.")
            return

        # ---- Date-shift pairs (Partner): Our Only D -> Supplier Only D+1..D+4 ----
        _pairs = find_date_shift_pairs(all_pending)
        if _pairs:
            st.markdown(f"#### 🔗 Date-shift pairs (Partner) — {len(_pairs)} pair(s) detected")
            _pair_rows = [{
                'Phone': display_phone(str(_ro.get('phone', '')).replace('.0', '')),
                'Amount': _ro.get('amount', ''),
                'Our Only date': _ro.get('date', ''),
                'Supplier Only date': _rs.get('date', ''),
                'Shift (days)': _dl,
            } for _ro, _rs, _dl in _pairs]
            st.dataframe(pd.DataFrame(_pair_rows), use_container_width=True, hide_index=True)
            st.caption("Rule: same phone + same amount, Our Only (D) → Supplier Only (D+1..D+4). "
                       "Will set: Our Only → '✅ Found — OK (date shift confirmed)', "
                       "Supplier Only → '✅ Found in our reports'.")
            if st.button(f"🔗 Verify all {len(_pairs)} pair(s) — one batch",
                         type="primary", key="pend_pairs_save"):
                sh_p = get_spreadsheet('partner')
                if sh_p is None:
                    st.error("❌ Google Sheets not connected. Click 🔄 Refresh.")
                else:
                    try:
                        ws_p = sh_p.worksheet('Transaction Details')
                        col_phones = _api_retry(lambda: ws_p.col_values(PHONE_COL))
                        cells, skipped = [], []
                        for _ro, _rs, _dl in _pairs:
                            for _rec, _status in (
                                    (_ro, "✅ Found — OK (date shift confirmed)"),
                                    (_rs, "✅ Found in our reports")):
                                _rn = _rec.get('_row')
                                _actual = col_phones[_rn - 1] if _rn and _rn - 1 < len(col_phones) else ''
                                if _rn and _norm_cmp(_actual) == _norm_cmp(_rec.get('phone')):
                                    cells.extend(_audit_cells(_rn, _status))
                                else:
                                    skipped.append(display_phone(str(_rec.get('phone', ''))))
                        if cells:
                            _api_retry(lambda: ws_p.update_cells(cells, value_input_option='RAW'))
                        _saved = {('partner', c.row) for c in cells}
                        st.session_state['pending_local'] = [
                            x for x in st.session_state.get('pending_local', [])
                            if (x.get('_op'), x.get('_row')) not in _saved]
                        _msg = f"✅ partner: verified {len({c.row for c in cells})} record(s) in {len(_pairs)} pair(s)"
                        if skipped:
                            _msg += (f" | skipped {len(skipped)} (rows moved — "
                                     f"click Refresh): {', '.join(skipped)}")
                        st.session_state['pend_last_result'] = [_msg]
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ {e}")
            st.markdown("---")

        STATUS_OPTIONS = ["⬜ Not checked", "✅ Found — OK (date shift confirmed)",
                          "✅ Found in our reports", "❌ Not found — investigate",
                          "🔵 Duplicate — refund issued to client"]

        bulk_mode = st.checkbox(
            "⚡ Bulk mode — set statuses in a table, save in one batch per operator",
            key="pend_bulk_mode")

        if bulk_mode:
            bulk_df = pd.DataFrame([{
                'Save': False,
                'Phone': display_phone(str(r.get('phone', '')).replace('.0', '')),
                'Date': r.get('date', ''),
                'Operator': r.get('_op', r.get('operator_tab', '')),
                'Category': r.get('category', ''),
                'Product': r.get('product', ''),
                'Amount': r.get('amount', ''),
                'New status': "⬜ Not checked",
                '_row': r.get('_row'),
            } for r in pending])
            edited = st.data_editor(
                bulk_df,
                column_config={
                    'Save': st.column_config.CheckboxColumn('Save'),
                    'New status': st.column_config.SelectboxColumn(
                        'New status', options=STATUS_OPTIONS),
                    '_row': None,
                },
                disabled=['Phone', 'Date', 'Operator', 'Category', 'Product', 'Amount'],
                use_container_width=True, hide_index=True, key="pend_bulk_editor")
            to_save = edited[(edited['Save']) & (edited['New status'] != "⬜ Not checked")]
            st.caption(f"Selected for save: {len(to_save)}")
            if st.button("💾 Save selected", type="primary", key="pend_bulk_save"):
                if len(to_save) == 0:
                    st.warning("Tick 'Save' and choose a status for at least one row.")
                else:
                    results = []
                    for op_name, grp in to_save.groupby('Operator'):
                        sh_op = get_spreadsheet(op_name)
                        if sh_op is None:
                            results.append(f"❌ {op_name}: Sheets not connected")
                            continue
                        try:
                            ws_op = sh_op.worksheet('Transaction Details')
                            col_phones = _api_retry(lambda: ws_op.col_values(PHONE_COL))
                            cells, skipped = [], []
                            for _, r in grp.iterrows():
                                if pd.isna(r['_row']):
                                    skipped.append(str(r['Phone']))
                                    continue
                                rn = int(r['_row'])
                                actual = col_phones[rn - 1] if rn - 1 < len(col_phones) else ''
                                if _norm_cmp(actual) == _norm_cmp(r['Phone']):
                                    cells.extend(_audit_cells(rn, r['New status']))
                                else:
                                    skipped.append(str(r['Phone']))
                            if cells:
                                _api_retry(lambda: ws_op.update_cells(
                                    cells, value_input_option='RAW'))
                            msg = f"✅ {op_name}: saved {len({c.row for c in cells})}"
                            if skipped:
                                msg += (f" | skipped {len(skipped)} "
                                        f"(rows moved — click Refresh): {', '.join(skipped)}")
                            results.append(msg)
                            saved = {(op_name, c.row) for c in cells}
                            st.session_state['pending_local'] = [
                                x for x in st.session_state.get('pending_local', [])
                                if (x.get('_op'), x.get('_row')) not in saved]
                        except Exception as e:
                            results.append(f"❌ {op_name}: {e}")
                    st.session_state['pend_last_result'] = results
                    st.rerun()
        else:
            if len(pending) > 50:
                st.info(f"Showing first 50 of {len(pending)} — use filter/search "
                        f"or ⚡ Bulk mode for mass updates.")
            for i, row in enumerate(pending[:50]):
                raw_phone = str(row.get('phone', '')).replace('.0', '')
                row_op = row.get('_op', row.get('operator_tab', 'partner'))
                sheet_row = row.get('_row')
                with st.expander(f"📱 {display_phone(raw_phone)} | {row.get('date', '')} | {str(row_op).upper()} | {row.get('category', '')}"):
                    c1, c2 = st.columns(2)
                    c1.write(f"**Product:** {row.get('product', '')}")
                    c1.write(f"**Amount:** {row.get('amount', '')} NIS")
                    c1.write(f"**Date:** {row.get('our_date', '') or row.get('sup_date', '')}")
                    c2.write("**What to check:**")
                    c2.info(row.get('check_instruction', ''))
                    _ukey = f"{row_op}_{sheet_row}"
                    new_status = st.selectbox("Update status:", STATUS_OPTIONS,
                                              key=f"pend_{_ukey}")
                    if st.button("Save", key=f"pend_save_{_ukey}"):
                        if new_status == "⬜ Not checked":
                            st.warning("Choose a status first.")
                        else:
                            sh = get_spreadsheet(row_op)
                            if sh is None:
                                st.error("❌ Google Sheets not connected. Click 🔄 Refresh.")
                            else:
                                ok, msg = update_verification_by_row(
                                    sh, sheet_row, raw_phone, new_status)
                                if ok:
                                    st.session_state['pending_local'] = [
                                        r for r in st.session_state.get('pending_local', [])
                                        if not (r.get('_op') == row_op
                                                and r.get('_row') == sheet_row)]
                                    st.session_state['pend_last_result'] = [
                                        f"✅ Saved {display_phone(raw_phone)}: {new_status}"]
                                    st.rerun()
                                else:
                                    st.error(f"❌ {msg}")


    # ============================================================
    # PAGE: VERIFIED
    # ============================================================
    elif page == "✅ Verified":
        render_header("Verified Transactions", "Completed verifications", [LOGO_PAYX])
        verified = load_verified()
        if not verified:
            st.info("No verified transactions yet.")
            return
        df = pd.DataFrame(verified)
        show_cols = [c for c in ['date','operator_tab','category','phone','product','amount','verified','verified_by','verified_at','check_instruction'] if c in df.columns]
        st.dataframe(df[show_cols], use_container_width=True, hide_index=True)
        st.success(f"✅ {len(verified)} transactions verified")

    # ============================================================
    # PAGE: INSTRUCTIONS
    # ============================================================
    elif page == "👥 Users":
        render_header("User Management", "Add employees, reset passwords, manage roles", [LOGO_PAYX])
        _my_role = str(st.session_state.get('auth_role', 'user')).strip().lower()
        if _my_role not in ('admin', 'owner'):
            st.error("Admins only.")
            return
        _me = str(st.session_state.get('auth_user', '')).strip().lower()

        if 'users_last_msg' in st.session_state:
            st.success(st.session_state.pop('users_last_msg'))

        users = load_users()

        def _role_of(u):
            return str(u.get('role', '')).strip().lower()

        # One-time bootstrap: while no active owner exists, an admin can claim it
        _owner_exists = any(_is_active(u) and _role_of(u) == 'owner' for u in users)
        if users and not _owner_exists and _my_role == 'admin':
            st.info("No owner is defined yet. Owner has full rights, "
                    "including granting and revoking admin.")
            if st.button("⭐ Become owner (one-time setup)", key="users_claim_owner"):
                _mine = next((u for u in users
                              if str(u.get('email', '')).strip().lower() == _me), None)
                if _mine:
                    _ws = _users_ws()
                    _api_retry(lambda: _ws.update_cell(
                        _mine['_row'], USERS_COLS.index('role') + 1, 'owner'))
                    load_users.clear()
                    st.session_state['auth_role'] = 'owner'
                    st.session_state['users_last_msg'] = f"⭐ {_me} is now owner."
                    st.rerun()

        if users:
            _udf = pd.DataFrame(users)
            _ucols = [c for c in ['email', 'role', 'active', 'created_by', 'created_at'] if c in _udf.columns]
            st.dataframe(_udf[_ucols], use_container_width=True, hide_index=True)
        else:
            st.info("No users yet. Create the first owner below.")

        _has_any_admin = any(_is_active(u) and _role_of(u) in ('admin', 'owner')
                             for u in users)

        st.markdown("#### ➕ Add user")
        with st.form("add_user_form"):
            _ne = st.text_input("Email")
            _np = st.text_input("Temporary password (min 8 characters)")
            if not _has_any_admin:
                _nr = "owner"
                st.caption("First user is created as **owner** (full rights).")
            elif _my_role == 'owner':
                _nr = st.selectbox("Role", ["user", "admin"])
            else:
                _nr = "user"
                st.caption("Admins can add **user** accounts only.")
            _sub = st.form_submit_button("Add user", type="primary")
        if _sub:
            _e = _ne.strip().lower()
            if not _e or '@' not in _e:
                st.error("Enter a valid email.")
            elif len(_np) < 8:
                st.error("Password must be at least 8 characters.")
            elif any(str(u.get('email', '')).strip().lower() == _e for u in users):
                st.error(f"User {_e} already exists. Use reset/activate below.")
            else:
                _ws = _users_ws()
                if _ws is None:
                    st.error("Google Sheets not connected.")
                else:
                    _api_retry(lambda: _ws.append_row(
                        [_e, _hash_pw(_np), _nr, '1',
                         st.session_state.get('auth_user', 'setup'), _now_il()]))
                    load_users.clear()
                    st.session_state['users_last_msg'] = f"✅ User {_e} added ({_nr})."
                    st.rerun()

        if users:
            st.markdown("#### 🔧 Manage user")
            if _my_role == 'owner':
                _managable = [u for u in users if _role_of(u) != 'owner'
                              and str(u.get('email', '')).strip().lower() != _me]
            else:
                _managable = [u for u in users if _role_of(u) == 'user'
                              and str(u.get('email', '')).strip().lower() != _me]
            _emails = [u.get('email', '') for u in _managable]
            if not _emails:
                st.info("No accounts available for you to manage."
                        + (" Admins manage user accounts only." if _my_role == 'admin' else ""))
            else:
                _sel = st.selectbox("Select user", _emails, key="users_manage_sel")
                _u = next((x for x in _managable if x.get('email', '') == _sel), None)
                if _u:
                    _c1, _c2, _c3 = st.columns(3)
                    with _c1:
                        if _is_active(_u):
                            if st.button("🚫 Deactivate", key="users_deact"):
                                _ws = _users_ws()
                                _api_retry(lambda: _ws.update_cell(
                                    _u['_row'], USERS_COLS.index('active') + 1, '0'))
                                load_users.clear()
                                st.session_state['users_last_msg'] = f"✅ {_sel} deactivated."
                                st.rerun()
                        else:
                            if st.button("✅ Activate", key="users_act"):
                                _ws = _users_ws()
                                _api_retry(lambda: _ws.update_cell(
                                    _u['_row'], USERS_COLS.index('active') + 1, '1'))
                                load_users.clear()
                                st.session_state['users_last_msg'] = f"✅ {_sel} activated."
                                st.rerun()
                    with _c2:
                        with st.form("reset_pw_form"):
                            _rp = st.text_input("New temporary password (min 8 chars)")
                            _rs = st.form_submit_button("🔑 Reset password")
                        if _rs:
                            if len(_rp) < 8:
                                st.error("Password must be at least 8 characters.")
                            else:
                                _ws = _users_ws()
                                _api_retry(lambda: _ws.update_cell(
                                    _u['_row'], USERS_COLS.index('pass_hash') + 1, _hash_pw(_rp)))
                                load_users.clear()
                                st.session_state['users_last_msg'] = f"✅ Password reset for {_sel}."
                                st.rerun()
                    with _c3:
                        if _my_role == 'owner':
                            if _role_of(_u) == 'user':
                                if st.button("⬆ Make admin", key="users_promote"):
                                    _ws = _users_ws()
                                    _api_retry(lambda: _ws.update_cell(
                                        _u['_row'], USERS_COLS.index('role') + 1, 'admin'))
                                    load_users.clear()
                                    st.session_state['users_last_msg'] = f"⬆ {_sel} is now admin."
                                    st.rerun()
                            elif _role_of(_u) == 'admin':
                                if st.button("⬇ Make user", key="users_demote"):
                                    _ws = _users_ws()
                                    _api_retry(lambda: _ws.update_cell(
                                        _u['_row'], USERS_COLS.index('role') + 1, 'user'))
                                    load_users.clear()
                                    st.session_state['users_last_msg'] = f"⬇ {_sel} is now user."
                                    st.rerun()
                        else:
                            st.caption("Role changes: owner only.")

    elif page == "ℹ️ Instructions":
        render_header("Instructions", "How to use the dashboard", [LOGO_PAYX])
        st.markdown("""
        ### 📋 Daily Process

        **Partner & 012Talk:**
        1. Upload Supplier `.xls` file
        2. Upload Partner EPRS `.csv` file
        3. Upload 012Talk EPRS `.csv` file
        4. Click **▶ Run Reconciliation**
        5. Review **Action Required** block — check each phone
        6. Click **💾 Save to Monthly History**

        **Pelephone:**
        1. Upload Supplier `.xlsx` file
        2. Upload Pelephone EPRS, GlobalSim EPRS, eSIM EPRS files
        3. Matching is done by Order Number = Transaction ID (exact match)
        4. eSIM transactions: expected difference is +2.67 NIS (supplier 7.67 vs our 5.00)

        **Cellcom:**
        1. Upload Supplier `.xlsx` file
        2. Upload Cellcom EPRS `.csv` file
        3. Fixed tariffs (15, 19, 49 NIS) — no price difference expected
        4. All other tariffs — supplier charges ~5 NIS less than our EUP
        5. "Unexplained Diff" should be 0.00 — if not, check anomaly rows

        ---

        ### 🔍 Verification Flow
        - **Action Required** block shows all phones needing verification
        - **What to check** column tells you exactly where to look
        - After checking, go to **⏳ Pending Verification** to update status
        - Verified items move to **✅ Verified**

        ---

        ### ⚠️ Important Rules
        - Always use original `.xls`/`.xlsx` supplier files — CSV truncates phone numbers
        - REFUND rows = credits from previous period
        - Late transactions (after 22:00) typically appear in NEXT day's supplier report
        - REWARD and REFUND_REWARD rows are automatically excluded

        ---

        ### 💾 About Saving
        - Each "Save to Monthly History" click **adds a new row** — it never overwrites old data
        - If you ran reconciliation twice for the same day, both rows are kept; the Monthly Summary shows the latest
        - If Google Sheets shows ⛔ in the sidebar — do NOT save until it's fixed
        """)


# ============================================================
# EXCEL EXPORTS
# ============================================================
def create_excel_report(result, report_date, tab_name):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def bd():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def H(cell, bg, fg='FFFFFF', sz=10):
        cell.font = Font(bold=True, color=fg, size=sz, name='Arial')
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = bd()
    def D(cell, bg='FFFFFF', align='left', fmt=None):
        cell.font = Font(size=9, name='Arial')
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = bd()
        if fmt: cell.number_format = fmt

    NAVY='1F3864'; LBLUE='DEEAF1'; RED='C00000'
    LRED='FFE0E0'; WHITE='FFFFFF'
    TEAL='00695C'; LTEAL='E0F2F1'

    t = result['totals']

    def write_sheet(ws, title_txt, df, hdr_bg, alt_bg, num_cols=None, ncols=None):
        nc = ncols or (len(df.columns) if len(df) > 0 else 8)
        ws.merge_cells(f'A1:{get_column_letter(nc)}1')
        ws['A1'].value = title_txt
        ws['A1'].font = Font(bold=True, color='FFFFFF', size=12, name='Arial')
        ws['A1'].fill = PatternFill('solid', start_color=hdr_bg)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].border = bd()
        ws.row_dimensions[1].height = 28
        if len(df) == 0: return
        for ci, col in enumerate(df.columns, 1):
            H(ws.cell(row=2, column=ci, value=col), hdr_bg)
        ws.row_dimensions[2].height = 30
        for ri, (_, row) in enumerate(df.iterrows()):
            r = ri + 3
            bg = alt_bg if ri%2==0 else WHITE
            for ci, col in enumerate(df.columns, 1):
                val = row[col]
                if pd.isna(val): val = ''
                cell = ws.cell(row=r, column=ci, value=val)
                is_num = num_cols and col in num_cols
                D(cell, bg, align='right' if is_num else 'left',
                  fmt='#,##0.00' if is_num else None)
            ws.row_dimensions[r].height = 15
        ws.freeze_panes = 'A3'

    ws_act = wb.create_sheet("Action Required")
    action_rows = []
    if len(result.get('sup_only', pd.DataFrame())) > 0:
        for _, r in result['sup_only'].iterrows():
            action_rows.append({
                'Side': 'Supplier Only',
                'Phone': r.get('Phone_Display', r.get('phone_norm','')),
                'Date': str(r.get('Sup_Date', r.get('sup_date',''))),
                'Product': r.get('Package', r.get('TOPUP_ITEM','')),
                'Amount (NIS)': r.get('CBD', r.get('TOPUP_PRICE',0)),
                'What to check': r.get('Check_Instruction',''),
                'Verified': '⬜ Not checked',
            })
    if len(result.get('our_only', pd.DataFrame())) > 0:
        for _, r in result['our_only'].iterrows():
            action_rows.append({
                'Side': 'Our Only',
                'Phone': r.get('Phone_Display',''),
                'Date': r.get('Date & Time',''),
                'Product': r.get('Product Name',''),
                'Amount (NIS)': r.get('End User Price',0),
                'What to check': r.get('Check_Instruction',''),
                'Verified': '⬜ Not checked',
            })
    action_df = pd.DataFrame(action_rows) if action_rows else pd.DataFrame(
        columns=['Side','Phone','Date','Product','Amount (NIS)','What to check','Verified'])
    write_sheet(ws_act, f"ACTION REQUIRED — {report_date} — {tab_name}",
                action_df, RED, LRED, {'Amount (NIS)'})
    if len(action_df) > 0:
        dv = DataValidation(type="list",
             formula1='"⬜ Not checked,✅ Found — OK,❌ Not found — investigate"',
             allow_blank=False, showDropDown=False)
        ver_col = list(action_df.columns).index('Verified') + 1
        vcl = get_column_letter(ver_col)
        dv.sqref = f"{vcl}3:{vcl}{len(action_df)+2}"
        ws_act.add_data_validation(dv)

    if len(result.get('matched', pd.DataFrame())) > 0:
        ws_m = wb.create_sheet("Matched")
        num_m = {c for c in result['matched'].columns if 'NIS' in c or 'Price' in c or 'Diff' in c}
        write_sheet(ws_m, f"MATCHED — {t.get('matched_count',0)} records", result['matched'],
                    NAVY, LBLUE, num_m)

    if len(result.get('refunds', pd.DataFrame())) > 0:
        ws_r = wb.create_sheet("Refunds")
        ref = result['refunds']
        show_cols = [c for c in ['Operator','Phone_Display','Date & Time','Product Name','End User Price'] if c in ref.columns]
        write_sheet(ws_r, f"REFUNDS — {t.get('refunds_count',0)} records",
                    ref[show_cols], TEAL, LTEAL, {'End User Price'})

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def create_monthly_excel(history, month_label):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"

    def bd():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)
    def H(cell, bg, fg='FFFFFF', sz=10):
        cell.font = Font(bold=True, color=fg, size=sz, name='Arial')
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = bd()
    def D(cell, bg='FFFFFF', fmt=None):
        cell.font = Font(size=9, name='Arial')
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd()
        if fmt: cell.number_format = fmt

    NAVY='1F3864'; LBLUE='DEEAF1'; GREEN='375623'; WHITE='FFFFFF'

    ws.merge_cells('A1:H1')
    ws['A1'].value = f"MONTHLY SUMMARY — {month_label}"
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14, name='Arial')
    ws['A1'].fill = PatternFill('solid', start_color=NAVY)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].border = bd()
    ws.row_dimensions[1].height = 36

    headers = ['Date', 'Operator', 'Matched', 'Supplier Total (NIS)',
               'Our EUP (NIS)', 'Diff (NIS)', 'Refunds (NIS)', 'Net Billed (NIS)']
    for ci, h in enumerate(headers, 1):
        H(ws.cell(row=2, column=ci, value=h), NAVY)
    ws.row_dimensions[2].height = 30

    for ri, rec in enumerate(history):
        r = ri + 3
        bg = LBLUE if ri%2==0 else WHITE
        vals = [rec.get('date',''), rec.get('operator_tab',''),
                rec.get('matched_count',0), rec.get('sup_cbd',0),
                rec.get('our_eup',0), rec.get('diff',0),
                rec.get('refunds_eup',0), rec.get('net_billed',0)]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            D(cell, bg, fmt='#,##0.00' if ci >= 4 else None)
        ws.row_dimensions[r].height = 18

    tr = len(history) + 3
    ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    for ci in range(1, 9):
        cell = ws.cell(row=tr, column=ci)
        if ci >= 4:
            cell.value = f'=SUM({get_column_letter(ci)}3:{get_column_letter(ci)}{tr-1})'
        H(cell, GREEN, sz=11)
        if ci >= 4: cell.number_format = '#,##0.00'
    ws.row_dimensions[tr].height = 24

    for ci, w in enumerate([14,12,10,20,18,16,16,18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


if __name__ == "__main__":
    main()
